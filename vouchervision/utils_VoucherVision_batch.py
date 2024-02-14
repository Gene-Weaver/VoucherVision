import openai
import os, sys, json, inspect, glob, tiktoken, shutil, yaml, torch, logging
import openpyxl
from openpyxl import Workbook, load_workbook
import google.generativeai as genai
import vertexai
from langchain_openai import AzureChatOpenAI
from transformers import TrOCRProcessor, VisionEncoderDecoderModel

currentdir = os.path.dirname(os.path.abspath(
    inspect.getfile(inspect.currentframe())))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)
parentdir = os.path.dirname(parentdir)
sys.path.append(parentdir)

from general_utils import get_cfg_from_full_path
# from embeddings_db import VoucherVisionEmbedding
from OCR_google_cloud_vision import OCRGoogle 

from vouchervision.LLM_OpenAI import OpenAIHandler
from vouchervision.LLM_GooglePalm2 import GooglePalm2Handler
from vouchervision.LLM_GoogleGemini import GoogleGeminiHandler
from vouchervision.LLM_MistralAI import MistralHandler
from vouchervision.LLM_local_cpu_MistralAI import LocalCPUMistralHandler
from vouchervision.LLM_local_MistralAI import LocalMistralHandler #call_llm_local_MistralAI_8x7b
from vouchervision.utils_LLM import remove_colons_and_double_apostrophes

# from LLM_Falcon import OCR_to_dict_Falcon
from prompt_catalog import PromptCatalog
from vouchervision.model_maps import ModelMaps

'''
* For the prefix_removal, the image names have 'MICH-V-' prior to the barcode, so that is used for matching
  but removed for output.
* There is also code active to replace the LLM-predicted "Catalog Number" with the correct number since it is known.
  The LLMs to usually assign the barcode to the correct field, but it's not needed since it is already known.
        - Look for ####################### Catalog Number pre-defined
'''


    
class VoucherVision():

    def __init__(self, cfg, logger, dir_home, path_custom_prompts, Project, Dirs):
        self.cfg = cfg
        self.logger = logger
        self.dir_home = dir_home
        self.path_custom_prompts = path_custom_prompts
        self.Project = Project
        self.Dirs = Dirs
        self.headers = None
        self.prompt_version = None

        # self.trOCR_model_version = "microsoft/trocr-large-handwritten"
        self.trOCR_model_version = "microsoft/trocr-base-handwritten"
        self.trOCR_processor = None
        self.trOCR_model = None

        self.set_API_keys()
        self.setup()


    def setup(self):
        self.logger.name = f'[Transcription]'
        self.logger.info(f'Setting up OCR and LLM')

        self.db_name = self.cfg['leafmachine']['project']['embeddings_database_name']
        self.path_domain_knowledge = self.cfg['leafmachine']['project']['path_to_domain_knowledge_xlsx']
        self.build_new_db = self.cfg['leafmachine']['project']['build_new_embeddings_database']

        self.continue_run_from_partial_xlsx = self.cfg['leafmachine']['project']['continue_run_from_partial_xlsx']

        self.prefix_removal = self.cfg['leafmachine']['project']['prefix_removal']
        self.suffix_removal = self.cfg['leafmachine']['project']['suffix_removal']
        self.catalog_numerical_only = self.cfg['leafmachine']['project']['catalog_numerical_only']

        self.prompt_version0 = self.cfg['leafmachine']['project']['prompt_version']
        self.use_domain_knowledge = self.cfg['leafmachine']['project']['use_domain_knowledge']

        self.catalog_name_options = ["Catalog Number", "catalog_number"]

        self.utility_headers = ["filename",
                                "WFO_override_OCR", "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_candidate_names","WFO_placement",
                                
                                "GEO_override_OCR", "GEO_method", "GEO_formatted_full_string", "GEO_decimal_lat",
                                "GEO_decimal_long","GEO_city", "GEO_county", "GEO_state",
                                "GEO_state_code", "GEO_country", "GEO_country_code", "GEO_continent",
                                
                                "tokens_in", "tokens_out", "path_to_crop","path_to_original","path_to_content","path_to_helper",]

        self.do_create_OCR_helper_image = self.cfg['leafmachine']['do_create_OCR_helper_image']

        self.map_prompt_versions()
        self.map_dir_labels()
        self.map_API_options()
        # self.init_embeddings()
        self.init_transcription_xlsx()
        self.init_trOCR_model()

        '''Logging'''
        self.logger.info(f'Transcribing dataset --- {self.dir_labels}')
        self.logger.info(f'Saving transcription batch to --- {self.path_transcription}')
        self.logger.info(f'Saving individual transcription files to --- {self.Dirs.transcription_ind}')
        self.logger.info(f'Starting transcription...')
        self.logger.info(f'     LLM MODEL --> {self.version_name}')
        self.logger.info(f'     Using Azure API --> {self.is_azure}')
        self.logger.info(f'     Model name passed to API --> {self.model_name}')
        self.logger.info(f'     API access token is found in PRIVATE_DATA.yaml --> {self.has_key}')

    def init_trOCR_model(self):
        lgr = logging.getLogger('transformers')
        lgr.setLevel(logging.ERROR)
        
        self.trOCR_processor = TrOCRProcessor.from_pretrained(self.trOCR_model_version)
        self.trOCR_model = VisionEncoderDecoderModel.from_pretrained(self.trOCR_model_version)
        
        # Check for GPU availability
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        self.trOCR_model.to(self.device)

    def map_API_options(self):
        self.chat_version = self.cfg['leafmachine']['LLM_version']

        # Get the required values from ModelMaps
        self.model_name = ModelMaps.get_version_mapping_cost(self.chat_version)
        self.is_azure = ModelMaps.get_version_mapping_is_azure(self.chat_version)
        self.has_key = ModelMaps.get_version_has_key(self.chat_version, self.has_key_openai, self.has_key_azure_openai, self.has_key_palm2, self.has_key_mistral)

        # Check if the version is supported
        if self.model_name is None:
            supported_LLMs = ", ".join(ModelMaps.get_models_gui_list())
            raise Exception(f"Unsupported LLM: {self.chat_version}. Requires one of: {supported_LLMs}")

        self.version_name = self.chat_version

    def map_prompt_versions(self):
        self.prompt_version_map = {
            "Version 1": "prompt_v1_verbose",
            "Version 1 No Domain Knowledge": "prompt_v1_verbose_noDomainKnowledge",
            "Version 2": "prompt_v2_json_rules",
            "Version 1 PaLM 2": 'prompt_v1_palm2',
            "Version 1 PaLM 2 No Domain Knowledge": 'prompt_v1_palm2_noDomainKnowledge', 
            "Version 2 PaLM 2": 'prompt_v2_palm2',
        }
        self.prompt_version = self.prompt_version_map.get(self.prompt_version0, self.path_custom_prompts)
        self.is_predefined_prompt = self.is_in_prompt_version_map(self.prompt_version)

    def is_in_prompt_version_map(self, value):
        return value in self.prompt_version_map.values()

    # def init_embeddings(self):
    #     if self.use_domain_knowledge:
    #         self.logger.info(f'*** USING DOMAIN KNOWLEDGE ***')
    #         self.logger.info(f'*** Initializing vector embeddings database ***')
    #         self.initialize_embeddings()
    #     else:
    #         self.Voucher_Vision_Embedding = None

    def map_dir_labels(self):
        if self.cfg['leafmachine']['use_RGB_label_images']:
            self.dir_labels = os.path.join(self.Dirs.save_per_annotation_class,'label')
        else:
            self.dir_labels = self.Dirs.save_original

        # Use glob to get all image paths in the directory
        self.img_paths = glob.glob(os.path.join(self.dir_labels, "*"))

    def load_rules_config(self):
        with open(self.path_custom_prompts, 'r') as stream:
            try:
                return yaml.safe_load(stream)
            except yaml.YAMLError as exc:
                print(exc)
                return None
            
    def generate_xlsx_headers(self):
        # Extract headers from the 'Dictionary' keys in the JSON template rules
        # xlsx_headers = list(self.rules_config_json['rules']["Dictionary"].keys())
        xlsx_headers = list(self.rules_config_json['rules'].keys())
        xlsx_headers = xlsx_headers + self.utility_headers
        return xlsx_headers

    def init_transcription_xlsx(self):
        self.HEADERS_v1_n22 = ["Catalog Number","Genus","Species","subspecies","variety","forma","Country","State","County","Locality Name","Min Elevation","Max Elevation","Elevation Units","Verbatim Coordinates","Datum","Cultivated","Habitat","Collectors","Collector Number","Verbatim Date","Date","End Date"] 
        self.HEADERS_v2_n26 = ["catalog_number","genus","species","subspecies","variety","forma","country","state","county","locality_name","min_elevation","max_elevation","elevation_units","verbatim_coordinates","decimal_coordinates","datum","cultivated","habitat","plant_description","collectors","collector_number","determined_by","multiple_names","verbatim_date","date","end_date"]
        self.HEADERS_v1_n22 = self.HEADERS_v1_n22 + self.utility_headers
        self.HEADERS_v2_n26 = self.HEADERS_v2_n26 + self.utility_headers
        # Initialize output file
        self.path_transcription = os.path.join(self.Dirs.transcription,"transcribed.xlsx")

        if self.prompt_version in ['prompt_v2_json_rules','prompt_v2_palm2']:
            self.headers = self.HEADERS_v2_n26
            self.headers_used = 'HEADERS_v2_n26'
        
        elif self.prompt_version in ['prompt_v1_verbose', 'prompt_v1_verbose_noDomainKnowledge','prompt_v1_palm2', 'prompt_v1_palm2_noDomainKnowledge']:
            self.headers = self.HEADERS_v1_n22
            self.headers_used = 'HEADERS_v1_n22'
        
        else:
            if not self.is_predefined_prompt:
                # Load the rules configuration
                self.rules_config_json = self.load_rules_config()
                # Generate the headers from the configuration
                self.headers = self.generate_xlsx_headers()
                # Set the headers used to the dynamically generated headers
                self.headers_used = 'CUSTOM'
            else:
                # If it's a predefined prompt, raise an exception as we don't have further instructions
                raise ValueError("Predefined prompt is not handled in this context.")

        self.create_or_load_excel_with_headers(os.path.join(self.Dirs.transcription,"transcribed.xlsx"), self.headers)

           
    def create_or_load_excel_with_headers(self, file_path, headers, show_head=False):
        output_dir_names = ['Archival_Components', 'Config_File', 'Cropped_Images', 'Logs', 'Original_Images', 'Transcription']
        self.completed_specimens = []

        # Check if the file exists and it's not None
        if self.continue_run_from_partial_xlsx is not None and os.path.isfile(self.continue_run_from_partial_xlsx):
            workbook = load_workbook(filename=self.continue_run_from_partial_xlsx)
            sheet = workbook.active
            show_head=True
            # Identify the 'path_to_crop' column
            try:
                path_to_crop_col = headers.index('path_to_crop') + 1
                path_to_original_col = headers.index('path_to_original') + 1
                path_to_content_col = headers.index('path_to_content') + 1
                path_to_helper_col = headers.index('path_to_helper') + 1
                # self.completed_specimens = list(sheet.iter_cols(min_col=path_to_crop_col, max_col=path_to_crop_col, values_only=True, min_row=2))
            except ValueError:
                print("'path_to_crop' not found in the header row.")

            
            path_to_crop = list(sheet.iter_cols(min_col=path_to_crop_col, max_col=path_to_crop_col, values_only=True, min_row=2))
            path_to_original = list(sheet.iter_cols(min_col=path_to_original_col, max_col=path_to_original_col, values_only=True, min_row=2))
            path_to_content = list(sheet.iter_cols(min_col=path_to_content_col, max_col=path_to_content_col, values_only=True, min_row=2))
            path_to_helper = list(sheet.iter_cols(min_col=path_to_helper_col, max_col=path_to_helper_col, values_only=True, min_row=2))
            others = [path_to_crop_col, path_to_original_col, path_to_content_col, path_to_helper_col]
            jsons = [path_to_content_col, path_to_helper_col]

            for cell in path_to_crop[0]:
                old_path = cell
                new_path = file_path
                for dir_name in output_dir_names:
                    if dir_name in old_path:
                        old_path_parts = old_path.split(dir_name)
                        new_path_parts = new_path.split('Transcription')
                        updated_path = new_path_parts[0] + dir_name + old_path_parts[1]
                        self.completed_specimens.append(os.path.basename(updated_path))
            print(f"{len(self.completed_specimens)} images are already completed")

            ### Copy the JSON files over
            for colu in jsons:
                cell = next(sheet.iter_rows(min_row=2, min_col=colu, max_col=colu))[0]
                old_path = cell.value
                new_path = file_path

                old_path_parts = old_path.split('Transcription')
                new_path_parts = new_path.split('Transcription')
                updated_path = new_path_parts[0] + 'Transcription' + old_path_parts[1]

                # Copy files
                old_dir = os.path.dirname(old_path)
                new_dir = os.path.dirname(updated_path)

                # Check if old_dir exists and it's a directory
                if os.path.exists(old_dir) and os.path.isdir(old_dir):
                    # Check if new_dir exists. If not, create it.
                    if not os.path.exists(new_dir):
                        os.makedirs(new_dir)

                    # Iterate through all files in old_dir and copy each to new_dir
                    for filename in os.listdir(old_dir):
                        shutil.copy2(os.path.join(old_dir, filename), new_dir) # copy2 preserves metadata

            ### Update the file names
            for colu in others:
                for row in sheet.iter_rows(min_row=2, min_col=colu, max_col=colu):
                    for cell in row:
                        old_path = cell.value
                        new_path = file_path
                        for dir_name in output_dir_names:
                            if dir_name in old_path:
                                old_path_parts = old_path.split(dir_name)
                                new_path_parts = new_path.split('Transcription')
                                updated_path = new_path_parts[0] + dir_name + old_path_parts[1]
                                cell.value = updated_path
            show_head=True

                
        else:
            # Create a new workbook and select the active worksheet
            workbook = Workbook()
            sheet = workbook.active

            # Write headers in the first row
            for i, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=i, value=header)
            self.completed_specimens = []
            
        # Save the workbook
        workbook.save(file_path)

        if show_head:
            print("continue_run_from_partial_xlsx:")
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                print(row)
                if i == 3:  # print the first 5 rows (0-indexed)
                    print("\n")
                    break



    def add_data_to_excel_from_response(self, path_transcription, response, WFO_record, GEO_record, filename_without_extension, path_to_crop, path_to_content, path_to_helper, nt_in, nt_out):
        geo_headers = ["GEO_override_OCR", "GEO_method", "GEO_formatted_full_string", "GEO_decimal_lat",
                       "GEO_decimal_long","GEO_city", "GEO_county", "GEO_state",
                       "GEO_state_code", "GEO_country", "GEO_country_code", "GEO_continent",]
        
        # WFO_candidate_names is separate, bc it may be type --> list
        wfo_headers = ["WFO_override_OCR", "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_placement"]

        wb = openpyxl.load_workbook(path_transcription)
        sheet = wb.active

        # find the next empty row
        next_row = sheet.max_row + 1

        if isinstance(response, str):
            try:
                response = json.loads(response)
            except json.JSONDecodeError:
                print(f"Failed to parse response: {response}")
                return

        # iterate over headers in the first row
        for i, header in enumerate(sheet[1], start=1):
            # check if header value is in response keys
            if (header.value in response) and (header.value not in self.catalog_name_options): ####################### Catalog Number pre-defined
                # check if the response value is a dictionary
                if isinstance(response[header.value], dict):
                    # if it is a dictionary, extract the 'value' field
                    cell_value = response[header.value].get('value', '')
                else:
                    # if it's not a dictionary, use it directly
                    cell_value = response[header.value]
                
                try:
                    # write the value to the cell
                    sheet.cell(row=next_row, column=i, value=cell_value)
                except:
                    sheet.cell(row=next_row, column=i, value=cell_value[0])

            elif header.value in self.catalog_name_options: 
                # if self.prefix_removal:
                #     filename_without_extension = filename_without_extension.replace(self.prefix_removal, "")
                # if self.suffix_removal:
                #     filename_without_extension = filename_without_extension.replace(self.suffix_removal, "")
                # if self.catalog_numerical_only:
                #     filename_without_extension = self.remove_non_numbers(filename_without_extension)
                sheet.cell(row=next_row, column=i, value=filename_without_extension)
            elif header.value == "path_to_crop":
                sheet.cell(row=next_row, column=i, value=path_to_crop)
            elif header.value == "path_to_original":
                if self.cfg['leafmachine']['use_RGB_label_images']:
                    fname = os.path.basename(path_to_crop)
                    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(path_to_crop))))
                    path_to_original = os.path.join(base, 'Original_Images', fname)
                    sheet.cell(row=next_row, column=i, value=path_to_original)
                else:
                    fname = os.path.basename(path_to_crop)
                    base = os.path.dirname(os.path.dirname(path_to_crop))
                    path_to_original = os.path.join(base, 'Original_Images', fname)
                    sheet.cell(row=next_row, column=i, value=path_to_original)
            elif header.value == "path_to_content":
                sheet.cell(row=next_row, column=i, value=path_to_content)
            elif header.value == "path_to_helper":
                sheet.cell(row=next_row, column=i, value=path_to_helper)
            elif header.value == "tokens_in":
                sheet.cell(row=next_row, column=i, value=nt_in)
            elif header.value == "tokens_out":
                sheet.cell(row=next_row, column=i, value=nt_out)
            elif header.value == "filename":
                sheet.cell(row=next_row, column=i, value=filename_without_extension)

            # "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_candidate_names","WFO_placement"
            elif header.value in wfo_headers:
                sheet.cell(row=next_row, column=i, value=WFO_record.get(header.value, ''))
            # elif header.value == "WFO_exact_match":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_exact_match",''))
            # elif header.value == "WFO_exact_match_name":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_exact_match_name",''))
            # elif header.value == "WFO_best_match":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_best_match",''))
            # elif header.value == "WFO_placement":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_placement",''))
            elif header.value == "WFO_candidate_names":
                candidate_names = WFO_record.get("WFO_candidate_names", '')
                # Check if candidate_names is a list and convert to a string if it is
                if isinstance(candidate_names, list):
                    candidate_names_str = '|'.join(candidate_names)
                else:
                    candidate_names_str = candidate_names
                sheet.cell(row=next_row, column=i, value=candidate_names_str)
            
            # "GEO_method", "GEO_formatted_full_string", "GEO_decimal_lat", "GEO_decimal_long",
            # "GEO_city", "GEO_county", "GEO_state", "GEO_state_code", "GEO_country", "GEO_country_code", "GEO_continent"
            elif header.value in geo_headers:
                sheet.cell(row=next_row, column=i, value=GEO_record.get(header.value, ''))

        # save the workbook
        wb.save(path_transcription)
    

    def has_API_key(self, val):
        if val != '':
            return True
        else:
            return False


    def set_API_keys(self):
        self.dir_home = os.path.dirname(os.path.dirname(__file__))
        self.path_cfg_private = os.path.join(self.dir_home, 'PRIVATE_DATA.yaml')
        self.cfg_private = get_cfg_from_full_path(self.path_cfg_private)

        self.has_key_openai = self.has_API_key(self.cfg_private['openai']['OPENAI_API_KEY'])

        self.has_key_azure_openai = self.has_API_key(self.cfg_private['openai_azure']['api_version'])
        
        self.has_key_palm2 = self.has_API_key(self.cfg_private['google_palm']['google_palm_api'])

        self.has_key_google_OCR = self.has_API_key(self.cfg_private['google_cloud']['path_json_file'])

        self.has_key_mistral = self.has_API_key(self.cfg_private['mistral']['mistral_key'])

        self.has_key_here = self.has_API_key(self.cfg_private['here']['api_key'])

        self.has_open_cage_geocode = self.has_API_key(self.cfg_private['open_cage_geocode']['api_key'])

        if self.has_key_openai:
            openai.api_key = self.cfg_private['openai']['OPENAI_API_KEY']
            os.environ["OPENAI_API_KEY"] = self.cfg_private['openai']['OPENAI_API_KEY']
            
        if self.has_key_azure_openai:
            # os.environ["OPENAI_API_KEY"] = self.cfg_private['openai_azure']['openai_api_key']
            self.llm = AzureChatOpenAI(
                deployment_name = 'gpt-35-turbo',#'gpt-35-turbo',
                openai_api_version = self.cfg_private['openai_azure']['api_version'],
                openai_api_key = self.cfg_private['openai_azure']['openai_api_key'],
                azure_endpoint = self.cfg_private['openai_azure']['openai_api_base'],
                # openai_api_base=self.cfg_private['openai_azure']['openai_api_base'],
                openai_organization = self.cfg_private['openai_azure']['openai_organization'],
                # openai_api_type = self.cfg_private['openai_azure']['openai_api_type']
            )
        
        # Enable this for all LLMs EXCEPT GOOGLE 
        name_check = self.cfg['leafmachine']['LLM_version'].lower().split(' ')
        if 'google' in name_check:
            pass
        else:
            if self.has_key_google_OCR:
                if os.path.exists(self.cfg_private['google_cloud']['path_json_file']):
                    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = self.cfg_private['google_cloud']['path_json_file']
                elif os.path.exists(self.cfg_private['google_cloud']['path_json_file_service_account2']):
                    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = self.cfg_private['google_cloud']['path_json_file_service_account2']
                else:
                    raise f"Google JSON API key file not found"

        if self.has_key_palm2:
            os.environ['PALM'] = self.cfg_private['google_palm']['google_palm_api']
            os.environ['PALM_PROJECT_ID'] = self.cfg_private['google_palm']['project_id']
            os.environ['PALM_LOCATION'] = self.cfg_private['google_palm']['location']
            os.environ['GOOGLE_API_KEY'] = self.cfg_private['google_palm']['google_palm_api']

            # os.environ["GOOGLE_SERVICE_ACCOUNT"] = self.cfg_private['google_cloud']['path_json_file_service_account']
            # vertexai.init(project='directed-curve-401601', location='us-central1')

            # genai.configure(api_key=os.environ['PALM'])

        if self.has_key_palm2:
            os.environ['MISTRAL_API_KEY'] = self.cfg_private['mistral']['mistral_key']

        if self.has_key_here:
            os.environ['here_app_id'] = self.cfg_private['here']['app_id']
            os.environ['here_api_key'] = self.cfg_private['here']['api_key']

        if self.has_open_cage_geocode:
            os.environ['open_cage_geocode'] = self.cfg_private['open_cage_geocode']['api_key']
        

        
    # def initialize_embeddings(self):
    #     '''Loading embedding search       __init__(self, db_name, path_domain_knowledge, logger, build_new_db=False, model_name="hkunlp/instructor-xl", device="cuda")'''
    #     self.Voucher_Vision_Embedding = VoucherVisionEmbedding(self.db_name, self.path_domain_knowledge, logger=self.logger, build_new_db=self.build_new_db)


    def clean_catalog_number(self, data, filename_without_extension):
        #Cleans up the catalog number in data if it's a dict
        
        def modify_catalog_key(catalog_key, filename_without_extension, data):
            # Helper function to apply modifications on catalog number
            if catalog_key not in data:
                new_data = {catalog_key: None}
                data = {**new_data, **data}

            if self.prefix_removal:
                filename_without_extension = filename_without_extension.replace(self.prefix_removal, "")
            if self.suffix_removal:
                filename_without_extension = filename_without_extension.replace(self.suffix_removal, "")
            if self.catalog_numerical_only:
                filename_without_extension = self.remove_non_numbers(data[catalog_key])
            data[catalog_key] = filename_without_extension
            return data
        
        if isinstance(data, dict):
            if self.headers_used == 'HEADERS_v1_n22':
                return modify_catalog_key("Catalog Number", filename_without_extension, data)
            elif self.headers_used in ['HEADERS_v2_n26', 'CUSTOM']:
                return modify_catalog_key("filename", filename_without_extension, data)
            else:
                raise ValueError("Invalid headers used.")
        else:
            raise TypeError("Data is not of type dict.")
        

    def write_json_to_file(self, filepath, data):
        '''Writes dictionary data to a JSON file.'''
        with open(filepath, 'w') as txt_file:
            if isinstance(data, dict):
                data = json.dumps(data, indent=4, sort_keys=False)
            txt_file.write(data)


    # def create_null_json(self):
    #     return {}
    

    def remove_non_numbers(self, s):
        return ''.join([char for char in s if char.isdigit()])
    

    def create_null_row(self, filename_without_extension, path_to_crop, path_to_content, path_to_helper):
        json_dict = {header: '' for header in self.headers} 
        for header, value in json_dict.items():
            if header == "path_to_crop":
                json_dict[header] = path_to_crop
            elif header == "path_to_original":
                fname = os.path.basename(path_to_crop)
                base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(path_to_crop))))
                path_to_original = os.path.join(base, 'Original_Images', fname)
                json_dict[header] = path_to_original
            elif header == "path_to_content":
                json_dict[header] = path_to_content
            elif header == "path_to_helper":
                json_dict[header] = path_to_helper
            elif header == "filename":
                json_dict[header] = filename_without_extension

            # "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_candidate_names","WFO_placement"
            elif header == "WFO_exact_match":
                json_dict[header] =''
            elif header == "WFO_exact_match_name":
                json_dict[header] = ''
            elif header == "WFO_best_match":
                json_dict[header] = ''
            elif header == "WFO_candidate_names":
                json_dict[header] = ''
            elif header == "WFO_placement":
                json_dict[header] = ''
        return json_dict
    

    ##################################################################################################################################
    ##################################################     OCR      ##################################################################
    ##################################################################################################################################
    def perform_OCR_and_save_results(self, image_index, jpg_file_path_OCR_helper, txt_file_path_OCR, txt_file_path_OCR_bounds):
        self.logger.info(f'Working on {image_index + 1}/{len(self.img_paths)} --- Starting OCR')
        # self.OCR - None

        ### Process_image() runs the OCR for text, handwriting, trOCR AND creates the overlay image
        ocr_google = OCRGoogle(self.path_to_crop, self.cfg, self.trOCR_model_version, self.trOCR_model, self.trOCR_processor, self.device)  
        ocr_google.process_image(self.do_create_OCR_helper_image, self.logger)
        self.OCR = ocr_google.OCR

        self.write_json_to_file(txt_file_path_OCR, ocr_google.OCR_JSON_to_file)
        
        self.logger.info(f'Working on {image_index + 1}/{len(self.img_paths)} --- Finished OCR')

        if len(self.OCR) > 0:
            ocr_google.overlay_image.save(jpg_file_path_OCR_helper)

            OCR_bounds = {}
            if ocr_google.hand_text_to_box_mapping is not None:
                OCR_bounds['OCR_bounds_handwritten'] = ocr_google.hand_text_to_box_mapping

            if ocr_google.normal_text_to_box_mapping is not None:
                OCR_bounds['OCR_bounds_printed'] = ocr_google.normal_text_to_box_mapping

            if ocr_google.trOCR_text_to_box_mapping is not None:
                OCR_bounds['OCR_bounds_trOCR'] = ocr_google.trOCR_text_to_box_mapping

            self.write_json_to_file(txt_file_path_OCR_bounds, OCR_bounds)
            self.logger.info(f'Working on {image_index + 1}/{len(self.img_paths)} --- Saved OCR Overlay Image')
        else:
            pass ########################################################################################################################### fix logic for no OCR

    ##################################################################################################################################
    #######################################################  LLM Switchboard  ########################################################
    ##################################################################################################################################
    def send_to_LLM(self, is_azure, progress_report, json_report, model_name):
        self.n_failed_LLM_calls = 0
        self.n_failed_OCR = 0

        final_JSON_response = None
        final_WFO_record = None
        final_GEO_record = None

        self.initialize_token_counters()
        self.update_progress_report_initial(progress_report)

        MODEL_NAME_FORMATTED = ModelMaps.get_API_name(model_name)
        name_parts = model_name.split("_")
        
        self.setup_JSON_dict_structure()
        
        llm_model = self.initialize_llm_model(self.logger, MODEL_NAME_FORMATTED, self.JSON_dict_structure, name_parts, is_azure, self.llm)

        prompts = []
        for i, path_to_crop in enumerate(self.img_paths):
            self.update_progress_report_batch(progress_report, i)

            if self.should_skip_specimen(path_to_crop):
                self.log_skipping_specimen(path_to_crop)
                continue

            paths = self.generate_paths(path_to_crop, i)
            self.path_to_crop = path_to_crop

            filename_without_extension, txt_file_path, txt_file_path_OCR, txt_file_path_OCR_bounds, jpg_file_path_OCR_helper = paths
            self.perform_OCR_and_save_results(i, jpg_file_path_OCR_helper, txt_file_path_OCR, txt_file_path_OCR_bounds)

            if not self.OCR:
                self.n_failed_OCR += 1
            else:
                # Format prompt
                prompt = self.setup_prompt()
                prompt = remove_colons_and_double_apostrophes(prompt)
                prompts.append(prompt)

        # Process prompts in batch
        if 'LOCAL' in name_parts and ('MISTRAL' in name_parts or 'MIXTRAL' in name_parts):
            batch_results = llm_model.call_llm_local_MistralAI(prompts)  # Assuming this method is updated to handle batch processing

        # Process each result from the batch
        for i, result in enumerate(batch_results):
            response_candidate, nt_in, nt_out, WFO_record, GEO_record = result.values()

            self.n_failed_LLM_calls += 1 if response_candidate is None else 0

            # Estimate n tokens returned
            self.logger.info(f'Prompt tokens IN --- {nt_in}')
            self.logger.info(f'Prompt tokens OUT --- {nt_out}')
            
            self.update_token_counters(nt_in, nt_out)

            final_JSON_response, final_WFO_record, final_GEO_record = self.update_final_response(response_candidate, WFO_record, GEO_record, paths, path_to_crop, nt_in, nt_out)

            self.log_completion_info(final_JSON_response)

            json_report.set_JSON(final_JSON_response, final_WFO_record, final_GEO_record)

        self.update_progress_report_final(progress_report)
        final_JSON_response = self.parse_final_json_response(final_JSON_response)
        return final_JSON_response, final_WFO_record, final_GEO_record, self.total_tokens_in, self.total_tokens_out

    

    ##################################################################################################################################
    ################################################## LLM Helper Funcs ##############################################################
    ##################################################################################################################################
    def initialize_llm_model(self, logger, model_name, JSON_dict_structure, name_parts, is_azure=None, llm_object=None):
        if 'LOCAL'in name_parts:
            if ('MIXTRAL' in name_parts) or ('MISTRAL' in name_parts):
                if 'CPU' in name_parts:
                    return LocalCPUMistralHandler(logger, model_name, JSON_dict_structure)
                else:
                    return LocalMistralHandler(logger, model_name, JSON_dict_structure)
        else:
            if 'PALM2' in name_parts:
                return GooglePalm2Handler(logger, model_name, JSON_dict_structure)
            elif 'GEMINI' in name_parts:
                return GoogleGeminiHandler(logger, model_name, JSON_dict_structure)
            elif 'MISTRAL' in name_parts and ('LOCAL' not in name_parts):
                return MistralHandler(logger, model_name, JSON_dict_structure)
            else:
                return OpenAIHandler(logger, model_name, JSON_dict_structure, is_azure, llm_object)

    def setup_prompt(self):
        Catalog = PromptCatalog()
        prompt, _ = Catalog.prompt_SLTP(self.path_custom_prompts, OCR=self.OCR)
        return prompt
    
    def setup_JSON_dict_structure(self):
        Catalog = PromptCatalog()
        _, self.JSON_dict_structure = Catalog.prompt_SLTP(self.path_custom_prompts, OCR='Text')
    

    def initialize_token_counters(self):
        self.total_tokens_in = 0
        self.total_tokens_out = 0


    def update_progress_report_initial(self, progress_report):
        if progress_report is not None:
            progress_report.set_n_batches(len(self.img_paths))


    def update_progress_report_batch(self, progress_report, batch_index):
        if progress_report is not None:
            progress_report.update_batch(f"Working on image {batch_index + 1} of {len(self.img_paths)}")


    def should_skip_specimen(self, path_to_crop):
        return os.path.basename(path_to_crop) in self.completed_specimens


    def log_skipping_specimen(self, path_to_crop):
        self.logger.info(f'[Skipping] specimen {os.path.basename(path_to_crop)} already processed')

    
    def update_token_counters(self, nt_in, nt_out):
        self.total_tokens_in += nt_in
        self.total_tokens_out += nt_out


    def update_final_response(self, response_candidate, WFO_record, GEO_record, paths, path_to_crop, nt_in, nt_out):
        filename_without_extension, txt_file_path, txt_file_path_OCR, txt_file_path_OCR_bounds, jpg_file_path_OCR_helper = paths
        # Saving the JSON and XLSX files with the response and updating the final JSON response
        if response_candidate is not None:
            final_JSON_response_updated = self.save_json_and_xlsx(response_candidate, WFO_record, GEO_record, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out)
            return final_JSON_response_updated, WFO_record, GEO_record
        else:
            final_JSON_response_updated = self.save_json_and_xlsx(response_candidate, WFO_record, GEO_record, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out)
            return final_JSON_response_updated, WFO_record, GEO_record


    def log_completion_info(self, final_JSON_response):
        self.logger.info(f'Formatted JSON\n{final_JSON_response}')
        self.logger.info(f'Finished API calls\n')


    def update_progress_report_final(self, progress_report):
        if progress_report is not None:
            progress_report.reset_batch("Batch Complete")


    def parse_final_json_response(self, final_JSON_response):
        try:
            return json.loads(final_JSON_response.strip('```').replace('json\n', '', 1).replace('json', '', 1))
        except:
            return final_JSON_response
    
    

    def generate_paths(self, path_to_crop, i):
        filename_without_extension = os.path.splitext(os.path.basename(path_to_crop))[0]
        txt_file_path = os.path.join(self.Dirs.transcription_ind, filename_without_extension + '.json')
        txt_file_path_OCR = os.path.join(self.Dirs.transcription_ind_OCR, filename_without_extension + '.json')
        txt_file_path_OCR_bounds = os.path.join(self.Dirs.transcription_ind_OCR_bounds, filename_without_extension + '.json')
        jpg_file_path_OCR_helper = os.path.join(self.Dirs.transcription_ind_OCR_helper, filename_without_extension + '.jpg')

        self.logger.info(f'Working on {i+1}/{len(self.img_paths)} --- {filename_without_extension}')

        return filename_without_extension, txt_file_path, txt_file_path_OCR, txt_file_path_OCR_bounds, jpg_file_path_OCR_helper


    def save_json_and_xlsx(self, response, WFO_record, GEO_record, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out):
        if response is None:
            response = self.JSON_dict_structure
            # Insert 'filename' as the first key
            response = {'filename': filename_without_extension, **{k: v for k, v in response.items() if k != 'filename'}}
            self.write_json_to_file(txt_file_path, response)

            # Then add the null info to the spreadsheet
            response_null = self.create_null_row(filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper)
            self.add_data_to_excel_from_response(self.path_transcription, response_null, WFO_record, GEO_record, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in=0, nt_out=0)
        
        ### Set completed JSON
        else:
            response = self.clean_catalog_number(response, filename_without_extension)
            self.write_json_to_file(txt_file_path, response)
            # add to the xlsx file
            self.add_data_to_excel_from_response(self.path_transcription, response, WFO_record, GEO_record, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out)
        return response
    

    def process_specimen_batch(self, progress_report, json_report, is_real_run=False):
        if not self.has_key:
            self.logger.error(f'No API key found for {self.version_name}')
            raise Exception(f"No API key found for {self.version_name}")

        try:
            if is_real_run:
                progress_report.update_overall(f"Transcribing Labels")

            final_json_response, final_WFO_record, final_GEO_record, total_tokens_in, total_tokens_out = self.send_to_LLM(self.is_azure, progress_report, json_report, self.model_name)
            
            return final_json_response, final_WFO_record, final_GEO_record, total_tokens_in, total_tokens_out

        except Exception as e:
            self.logger.error(f"LLM call failed in process_specimen_batch: {e}")
            if progress_report is not None:
                progress_report.reset_batch(f"Batch Failed")
            self.close_logger_handlers()
            raise


    def close_logger_handlers(self):
        for handler in self.logger.handlers[:]:
            handler.close()
            self.logger.removeHandler(handler)


    def process_specimen_batch_OCR_test(self, path_to_crop):
        for img_filename in os.listdir(path_to_crop):
            img_path = os.path.join(path_to_crop, img_filename)
        self.OCR, self.bounds, self.text_to_box_mapping = detect_text(img_path)



def space_saver(cfg, Dirs, logger):
    dir_out = cfg['leafmachine']['project']['dir_output']
    run_name = Dirs.run_name

    path_project = os.path.join(dir_out, run_name)

    if cfg['leafmachine']['project']['delete_temps_keep_VVE']:
        logger.name = '[DELETE TEMP FILES]'
        logger.info("Deleting temporary files. Keeping files required for VoucherVisionEditor.")
        delete_dirs = ['Archival_Components', 'Config_File']
        for d in delete_dirs:
            path_delete = os.path.join(path_project, d)
            if os.path.exists(path_delete):
                shutil.rmtree(path_delete)

    elif cfg['leafmachine']['project']['delete_all_temps']:
        logger.name = '[DELETE TEMP FILES]'
        logger.info("Deleting ALL temporary files!")
        delete_dirs = ['Archival_Components', 'Config_File', 'Original_Images', 'Cropped_Images']
        for d in delete_dirs:
            path_delete = os.path.join(path_project, d)
            if os.path.exists(path_delete):
                shutil.rmtree(path_delete)

        # Delete the transctiption folder, but keep the xlsx
        transcription_path = os.path.join(path_project, 'Transcription')
        if os.path.exists(transcription_path):
            for item in os.listdir(transcription_path):
                item_path = os.path.join(transcription_path, item)
                if os.path.isdir(item_path):  # if the item is a directory
                    if os.path.exists(item_path):
                        shutil.rmtree(item_path)  # delete the directory
