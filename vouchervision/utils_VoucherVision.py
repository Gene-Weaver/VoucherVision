import concurrent
import openai
from tqdm import tqdm
import os, json, glob, shutil, yaml, torch, logging, gc, traceback
import openpyxl
from openpyxl import Workbook, load_workbook
import vertexai
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
from langchain_openai import AzureChatOpenAI
from google.oauth2 import service_account
from transformers import AutoTokenizer, AutoModel
import pandas as pd

import queue
import threading

# from vouchervision.LLM_Hyperbolic import HyperbolicHandler
from vouchervision.prompt_catalog import PromptCatalog
from vouchervision.model_maps import ModelMaps
from vouchervision.general_utils import get_cfg_from_full_path
from vouchervision.OCR_google_cloud_vision import OCREngine 

'''
* For the prefix_removal, the image names have 'MICH-V-' prior to the barcode, so that is used for matching
  but removed for output.
* There is also code active to replace the LLM-predicted "Catalog Number" with the correct number since it is known.
  The LLMs to usually assign the barcode to the correct field, but it's not needed since it is already known.
        - Look for ####################### Catalog Number pre-defined
'''

def process_single_image_worker(job_args):
    """
    A stateless worker function to process a single image.
    This function is designed to be thread-safe as it does not modify shared state.
    """
    # Unpack all arguments from the job tuple
    (
        i, path_to_crop, total_images, llm_model, ocr_engine,
        vv_instance, model_name_formatted, name_parts
    ) = job_args

    # Get a logger instance
    logger = logging.getLogger(__name__)
    worker_id = threading.current_thread().name
    logger.info(f'[{worker_id}] Working on {i+1}/{total_images} --- {os.path.basename(path_to_crop)}')

    # --- 1. Generate Paths ---
    paths = vv_instance.generate_paths(path_to_crop, i)
    (
        filename_without_extension, txt_file_path, txt_file_path_OCR,
        txt_file_path_OCR_bounds, jpg_file_path_OCR_helper,
        json_file_path_wiki, txt_file_path_ind_prompt
    ) = paths

    # --- 2. Perform OCR ---
    ocr_text = ""
    ocr_cost, ocr_tokens_in, ocr_tokens_out, ocr_cost_in, ocr_cost_out = 0, 0, 0, 0, 0
    ocr_method = ""
    ocr_failed = False
    try:
        # Perform OCR using the provided engine instance
        ocr_engine.process_image(vv_instance.do_create_OCR_helper_image, path_to_crop, logger)
        ocr_text = ocr_engine.OCR
        
        # Capture OCR stats from the engine instance used by this worker
        ocr_cost = ocr_engine.cost
        ocr_tokens_in = ocr_engine.tokens_in
        ocr_tokens_out = ocr_engine.tokens_out
        ocr_cost_in = ocr_engine.cost_in
        ocr_cost_out = ocr_engine.cost_out
        ocr_method = str(ocr_engine.ocr_method)

        if not ocr_text:
            ocr_failed = True
            logger.warning(f"[{worker_id}] OCR resulted in empty text for {os.path.basename(path_to_crop)}")
        else:
             # Save OCR helper files
            vv_instance.write_json_to_file(txt_file_path_OCR, ocr_engine.OCR_JSON_to_file)
            if ocr_engine.overlay_image:
                 ocr_engine.overlay_image.save(jpg_file_path_OCR_helper)

    except Exception as e:
        ocr_failed = True
        logger.error(f"[{worker_id}] OCR processing failed for {os.path.basename(path_to_crop)}: {e}")
        logger.error(traceback.format_exc())

    # --- 3. Call LLM (if OCR succeeded) ---
    response_candidate = None
    nt_in, nt_out = 0, 0
    WFO_record, GEO_record, usage_report = None, None, None
    llm_failed = False

    if not ocr_failed:
        try:
            # Generate prompt by passing ocr_text directly
            prompt = vv_instance.setup_prompt(ocr_text)

            # Call the appropriate LLM based on model name
            # Note: json_report is passed as None since we can't update a GUI from a worker thread.
            if 'PALM2' in name_parts:
                response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_GooglePalm2(prompt, None, paths)
            elif 'GEMINI' in name_parts:
                response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_GoogleGemini(prompt, None, paths)
            elif 'MISTRAL' in name_parts and ('LOCAL' not in name_parts):
                response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_MistralAI(prompt, None, paths)
            elif 'Hyperbolic' in name_parts:
                response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_Hyperbolic(prompt, None, paths)
            else: # Fallback to OpenAI
                response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_OpenAI(prompt, None, paths)

            if response_candidate is None:
                llm_failed = True
                logger.error(f"[{worker_id}] LLM call failed for {os.path.basename(path_to_crop)}")

        except Exception as e:
            llm_failed = True
            logger.error(f"[{worker_id}] LLM call failed with exception for {os.path.basename(path_to_crop)}: {e}")
            logger.error(traceback.format_exc())

    # --- 4. Package and Return All Results ---
    result_package = {
        'index': i,
        'path_to_crop': path_to_crop,
        'paths': paths,
        'ocr_failed': ocr_failed,
        'llm_failed': llm_failed,
        'response_candidate': response_candidate,
        'nt_in': nt_in,
        'nt_out': nt_out,
        'WFO_record': WFO_record,
        'GEO_record': GEO_record,
        'usage_report': usage_report,
        'ocr_cost': ocr_cost,
        'ocr_tokens_in': ocr_tokens_in,
        'ocr_tokens_out': ocr_tokens_out,
        'ocr_cost_in': ocr_cost_in,
        'ocr_cost_out': ocr_cost_out,
        'ocr_method': ocr_method
    }
    return result_package
    
class VoucherVision():

    def __init__(self, cfg, logger, dir_home, path_custom_prompts, Project, Dirs, is_hf, config_vals_for_permutation=None, skip_API_keys=False):
        self.cfg = cfg
        self.logger = logger
        self.dir_home = dir_home
        self.path_custom_prompts = path_custom_prompts
        self.Project = Project
        self.Dirs = Dirs
        self.headers = None
        self.prompt_version = None
        self.is_hf = is_hf
        self.skip_API_keys = skip_API_keys

        self.OCR_cost = 0.0
        self.OCR_tokens_in = 0
        self.OCR_tokens_out = 0
        self.OCR_cost_in = 0.0
        self.OCR_cost_out = 0.0
        self.ocr_method = ""

        ### config_vals_for_permutation allows you to set the starting temp, top_k, top_p, seed....
        self.config_vals_for_permutation = config_vals_for_permutation

        # self.trOCR_model_version = "microsoft/trocr-large-handwritten"
        # self.trOCR_model_version = "microsoft/trocr-base-handwritten"
        # self.trOCR_model_version = "dh-unibe/trocr-medieval-escriptmask" # NOPE
        # self.trOCR_model_version = "dh-unibe/trocr-kurrent" # NOPE
        # self.trOCR_model_version = "DunnBC22/trocr-base-handwritten-OCR-handwriting_recognition_v2" # NOPE
        self.trOCR_processor = None
        self.trOCR_model = None

        # if not self.skip_API_keys: # This is for VVGO
        #     self.set_API_keys()
        # else:
        #     self.has_key_openai = False
        #     self.has_key_azure_openai = False
        #     self.has_key_google_application_credentials = True
        #     self.has_key_mistral = False
        #     self.has_key_hyperbolic = False
        self.set_API_keys()
        self.setup()


    def setup(self):
        if not self.skip_API_keys: 
            self.logger.name = f'[Transcription]'
            self.logger.info(f'Setting up OCR and LLM')

        self.trOCR_model_version = self.cfg['leafmachine']['project']['trOCR_model_path']

        # self.db_name = self.cfg['leafmachine']['project']['embeddings_database_name']
        # self.path_domain_knowledge = self.cfg['leafmachine']['project']['path_to_domain_knowledge_xlsx']
        # self.build_new_db = self.cfg['leafmachine']['project']['build_new_embeddings_database']

        self.continue_run_from_partial_xlsx = self.cfg['leafmachine']['project']['continue_run_from_partial_xlsx']

        self.prefix_removal = self.cfg['leafmachine']['project']['prefix_removal']
        self.suffix_removal = self.cfg['leafmachine']['project']['suffix_removal']
        self.catalog_numerical_only = self.cfg['leafmachine']['project']['catalog_numerical_only']

        self.prompt_version0 = self.cfg['leafmachine']['project']['prompt_version']
        self.use_domain_knowledge = self.cfg['leafmachine']['project']['use_domain_knowledge']

        self.catalog_name_options = ["Catalog Number", "catalog_number", "catalogNumber"]

        self.geo_headers = ["GEO_override_OCR", "GEO_method", "GEO_formatted_full_string", "GEO_decimal_lat",
                       "GEO_decimal_long","GEO_city", "GEO_county", "GEO_state",
                       "GEO_state_code", "GEO_country", "GEO_country_code", "GEO_continent",]
        
        self.usage_headers = ["current_time", "inference_time_s", "tool_time_s","max_cpu", "max_ram_gb", "n_gpus", "max_gpu_load", "max_gpu_vram_gb","total_gpu_vram_gb","capability_score",]
        
        self.wfo_headers = ["WFO_override_OCR", "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_candidate_names","WFO_placement"]
        self.wfo_headers_no_lists = ["WFO_override_OCR", "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_placement"]
        
        self.utility_headers = ["filename"] + self.wfo_headers + self.geo_headers + self.usage_headers + ["run_name", "prompt", "LLM", "tokens_in", "tokens_out", "LM2_collage", "OCR_method", "OCR_double", "OCR_trOCR", "path_to_crop","path_to_original","path_to_content","path_to_helper",]
                                # "WFO_override_OCR", "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_candidate_names","WFO_placement",
                                
                                # "GEO_override_OCR", "GEO_method", "GEO_formatted_full_string", "GEO_decimal_lat",
                                # "GEO_decimal_long","GEO_city", "GEO_county", "GEO_state",
                                # "GEO_state_code", "GEO_country", "GEO_country_code", "GEO_continent",
                                
                                # "tokens_in", "tokens_out", "path_to_crop","path_to_original","path_to_content","path_to_helper",]
        
        # WFO_candidate_names is separate, bc it may be type --> list

        self.do_create_OCR_helper_image = self.cfg['leafmachine']['do_create_OCR_helper_image']

        if not self.skip_API_keys: 
            self.map_prompt_versions()
            self.map_dir_labels()
        self.map_API_options()
        
        if not self.skip_API_keys: 
            # self.init_embeddings()
            self.init_transcription_xlsx()
            self.init_trOCR_model()

            '''Logging'''
            self.logger.info(f'Transcribing dataset --- {self.dir_labels}')
            self.logger.info(f'Saving transcription batch to --- {self.path_transcription}')
            self.logger.info(f'Saving individual transcription files to --- {self.Dirs.transcription_ind}')
            self.logger.info(f'Starting transcription...')
            self.logger.info(f'     LLM MODEL --> {self.version_name}')
            self.logger.info(f'     Using Azure API --> {self.is_azure}')
            self.logger.info(f'     Model name passed to API --> {self.model_name}')
            self.logger.info(f'     API access token is found in PRIVATE_DATA.yaml --> {self.has_key}')


    def init_trOCR_model(self):
        lgr = logging.getLogger('transformers')
        lgr.setLevel(logging.ERROR)
        
        self.trOCR_processor = TrOCRProcessor.from_pretrained("microsoft/trocr-base-handwritten") # usually just the "microsoft/trocr-base-handwritten"
        self.trOCR_model = VisionEncoderDecoderModel.from_pretrained(self.trOCR_model_version) # This matches the model
        
        # Check for GPU availability
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        self.trOCR_model.to(self.device)


    def map_API_options(self):
        self.chat_version = self.cfg['leafmachine']['LLM_version']

        # Get the required values from ModelMaps
        self.model_name = ModelMaps.get_version_mapping_cost(self.chat_version)
        self.is_azure = ModelMaps.get_version_mapping_is_azure(self.chat_version)
        self.has_key = ModelMaps.get_version_has_key(self.chat_version, self.has_key_openai, self.has_key_azure_openai, self.has_key_google_application_credentials, self.has_key_mistral, self.has_key_hyperbolic)

        # Check if the version is supported
        if self.model_name is None:
            supported_LLMs = ", ".join(ModelMaps.get_models_gui_list())
            raise Exception(f"Unsupported LLM: {self.chat_version}. Requires one of: {supported_LLMs}")

        self.version_name = self.chat_version


    def map_prompt_versions(self):
        self.prompt_version_map = {
            "Version 1": "prompt_v1_verbose",
        }
        self.prompt_version = self.prompt_version_map.get(self.prompt_version0, self.path_custom_prompts)
        self.is_predefined_prompt = self.is_in_prompt_version_map(self.prompt_version)


    def is_in_prompt_version_map(self, value):
        return value in self.prompt_version_map.values()


    def map_dir_labels(self):
        if self.cfg['leafmachine']['use_RGB_label_images'] in [1,2]:
            self.dir_labels = os.path.join(self.Dirs.save_per_annotation_class,'label')
        else:
            self.dir_labels = self.Dirs.save_original

        # Use glob to get all image paths in the directory
        self.img_paths = glob.glob(os.path.join(self.dir_labels, "*"))


    def load_rules_config(self):
        with open(self.path_custom_prompts, 'r') as stream:
            try:
                return yaml.safe_load(stream)
            except yaml.YAMLError as exc:
                print(exc)
                return None
            

    def generate_xlsx_headers(self):
        # Extract headers from the 'Dictionary' keys in the JSON template rules
        # xlsx_headers = list(self.rules_config_json['rules']["Dictionary"].keys())
        xlsx_headers = list(self.rules_config_json['rules'].keys())
        xlsx_headers = xlsx_headers + self.utility_headers
        return xlsx_headers


    def init_transcription_xlsx(self):
        # Initialize output file
        self.path_transcription = os.path.join(self.Dirs.transcription,"transcribed.xlsx")
        
        # else:
        if not self.is_predefined_prompt:
            # Load the rules configuration
            self.rules_config_json = self.load_rules_config()
            # Generate the headers from the configuration
            self.headers = self.generate_xlsx_headers()
            # Set the headers used to the dynamically generated headers
            self.headers_used = 'CUSTOM'
        else:
            # If it's a predefined prompt, raise an exception as we don't have further instructions
            raise ValueError("Predefined prompt is not handled in this context.")

        self.create_or_load_excel_with_headers(os.path.join(self.Dirs.transcription,"transcribed.xlsx"), self.headers)

           
    def create_or_load_excel_with_headers(self, file_path, headers, show_head=False):
        output_dir_names = ['Archival_Components', 'Config_File', 'Cropped_Images', 'Logs', 'Original_Images', 'Transcription']
        self.completed_specimens = []

        # Check if the file exists and it's not None
        if self.continue_run_from_partial_xlsx is not None and os.path.isfile(self.continue_run_from_partial_xlsx):
            workbook = load_workbook(filename=self.continue_run_from_partial_xlsx)
            sheet = workbook.active
            show_head=True
            # Identify the 'path_to_crop' column
            try:
                path_to_crop_col = headers.index('path_to_crop') + 1
                path_to_original_col = headers.index('path_to_original') + 1
                path_to_content_col = headers.index('path_to_content') + 1
                path_to_helper_col = headers.index('path_to_helper') + 1
                # self.completed_specimens = list(sheet.iter_cols(min_col=path_to_crop_col, max_col=path_to_crop_col, values_only=True, min_row=2))
            except ValueError:
                print("'path_to_crop' not found in the header row.")

            path_to_crop = list(sheet.iter_cols(min_col=path_to_crop_col, max_col=path_to_crop_col, values_only=True, min_row=2))
            path_to_original = list(sheet.iter_cols(min_col=path_to_original_col, max_col=path_to_original_col, values_only=True, min_row=2))
            path_to_content = list(sheet.iter_cols(min_col=path_to_content_col, max_col=path_to_content_col, values_only=True, min_row=2))
            path_to_helper = list(sheet.iter_cols(min_col=path_to_helper_col, max_col=path_to_helper_col, values_only=True, min_row=2))
            others = [path_to_crop_col, path_to_original_col, path_to_content_col, path_to_helper_col]
            jsons = [path_to_content_col, path_to_helper_col]

            for cell in path_to_crop[0]:
                old_path = cell
                new_path = file_path
                for dir_name in output_dir_names:
                    if dir_name in old_path:
                        old_path_parts = old_path.split(dir_name)
                        new_path_parts = new_path.split('Transcription')
                        updated_path = new_path_parts[0] + dir_name + old_path_parts[1]
                        self.completed_specimens.append(os.path.basename(updated_path))
            print(f"{len(self.completed_specimens)} images are already completed")

            ### Copy the JSON files over
            for colu in jsons:
                cell = next(sheet.iter_rows(min_row=2, min_col=colu, max_col=colu))[0]
                old_path = cell.value
                new_path = file_path

                old_path_parts = old_path.split('Transcription')
                new_path_parts = new_path.split('Transcription')
                updated_path = new_path_parts[0] + 'Transcription' + old_path_parts[1]

                # Copy files
                old_dir = os.path.dirname(old_path)
                new_dir = os.path.dirname(updated_path)

                # Check if old_dir exists and it's a directory
                if os.path.exists(old_dir) and os.path.isdir(old_dir):
                    # Check if new_dir exists. If not, create it.
                    if not os.path.exists(new_dir):
                        os.makedirs(new_dir)

                    # Iterate through all files in old_dir and copy each to new_dir
                    for filename in os.listdir(old_dir):
                        shutil.copy2(os.path.join(old_dir, filename), new_dir) # copy2 preserves metadata

            ### Update the file names
            for colu in others:
                for row in sheet.iter_rows(min_row=2, min_col=colu, max_col=colu):
                    for cell in row:
                        old_path = cell.value
                        new_path = file_path
                        for dir_name in output_dir_names:
                            if dir_name in old_path:
                                old_path_parts = old_path.split(dir_name)
                                new_path_parts = new_path.split('Transcription')
                                updated_path = new_path_parts[0] + dir_name + old_path_parts[1]
                                cell.value = updated_path
            show_head=True

                
        else:
            # Create a new workbook and select the active worksheet
            workbook = Workbook()
            sheet = workbook.active

            # Write headers in the first row
            for i, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=i, value=header)
            self.completed_specimens = []
            
        # Save the workbook
        workbook.save(file_path)

        if show_head:
            print("continue_run_from_partial_xlsx:")
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                print(row)
                if i == 3:  # print the first 5 rows (0-indexed)
                    print("\n")
                    break


    def add_data_to_excel_from_response(self, Dirs, path_transcription, response, WFO_record, GEO_record, usage_report, 
                                        MODEL_NAME_FORMATTED, filename_without_extension, path_to_crop, path_to_content, path_to_helper, nt_in, nt_out):
        

        wb = openpyxl.load_workbook(path_transcription)
        sheet = wb.active

        # find the next empty row
        next_row = sheet.max_row + 1

        if isinstance(response, str):
            try:
                response = json.loads(response)
            except json.JSONDecodeError:
                print(f"Failed to parse response: {response}")
                return

        # iterate over headers in the first row
        for i, header in enumerate(sheet[1], start=1):
            # check if header value is in response keys
            if (header.value in response) and (header.value not in self.catalog_name_options): ####################### Catalog Number pre-defined
                # check if the response value is a dictionary
                if isinstance(response[header.value], dict):
                    # if it is a dictionary, extract the 'value' field
                    cell_value = response[header.value].get('value', '')
                else:
                    # if it's not a dictionary, use it directly
                    cell_value = response[header.value]
                
                try:
                    # write the value to the cell
                    sheet.cell(row=next_row, column=i, value=cell_value)
                except:
                    sheet.cell(row=next_row, column=i, value=cell_value[0])

            elif header.value in self.catalog_name_options: 
                # if self.prefix_removal:
                #     filename_without_extension = filename_without_extension.replace(self.prefix_removal, "")
                # if self.suffix_removal:
                #     filename_without_extension = filename_without_extension.replace(self.suffix_removal, "")
                # if self.catalog_numerical_only:
                #     filename_without_extension = self.remove_non_numbers(filename_without_extension)
                sheet.cell(row=next_row, column=i, value=filename_without_extension)
            elif header.value == "path_to_crop":
                sheet.cell(row=next_row, column=i, value=path_to_crop)
            elif header.value == "path_to_original":
                if self.cfg['leafmachine']['use_RGB_label_images'] in [1,2]:
                    fname = os.path.basename(path_to_crop)
                    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(path_to_crop))))
                    path_to_original = os.path.join(base, 'Original_Images', fname)
                    sheet.cell(row=next_row, column=i, value=path_to_original)
                else:
                    fname = os.path.basename(path_to_crop)
                    base = os.path.dirname(os.path.dirname(path_to_crop))
                    path_to_original = os.path.join(base, 'Original_Images', fname)
                    sheet.cell(row=next_row, column=i, value=path_to_original)
            elif header.value == "path_to_content":
                sheet.cell(row=next_row, column=i, value=path_to_content)
            elif header.value == "path_to_helper":
                sheet.cell(row=next_row, column=i, value=path_to_helper)
            elif header.value == "tokens_in":
                sheet.cell(row=next_row, column=i, value=nt_in)
            elif header.value == "tokens_out":
                sheet.cell(row=next_row, column=i, value=nt_out)
            elif header.value == "filename":
                sheet.cell(row=next_row, column=i, value=filename_without_extension)
            elif header.value == "prompt":
                sheet.cell(row=next_row, column=i, value=os.path.basename(self.path_custom_prompts))
            elif header.value == "run_name":
                sheet.cell(row=next_row, column=i, value=Dirs.run_name)
            elif header.value == "LM2_collage":
                sheet.cell(row=next_row, column=i, value=self.cfg['leafmachine']['use_RGB_label_images'])
            elif header.value == "OCR_method":
                value_to_insert = self.cfg['leafmachine']['project']['OCR_option']
                if isinstance(value_to_insert, list):
                    value_to_insert = '|'.join(map(str, value_to_insert))
                sheet.cell(row=next_row, column=i, value=value_to_insert)
            elif header.value == "OCR_double":
                sheet.cell(row=next_row, column=i, value=self.cfg['leafmachine']['project']['double_OCR'])
            elif header.value == "OCR_trOCR":
                sheet.cell(row=next_row, column=i, value=self.cfg['leafmachine']['project']['do_use_trOCR'])
            # "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_candidate_names","WFO_placement"
            elif header.value in self.wfo_headers_no_lists:
                sheet.cell(row=next_row, column=i, value=WFO_record.get(header.value, ''))
            # elif header.value == "WFO_exact_match":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_exact_match",''))
            # elif header.value == "WFO_exact_match_name":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_exact_match_name",''))
            # elif header.value == "WFO_best_match":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_best_match",''))
            # elif header.value == "WFO_placement":
            #     sheet.cell(row=next_row, column=i, value= WFO_record.get("WFO_placement",''))
            elif header.value == "WFO_candidate_names":
                candidate_names = WFO_record.get("WFO_candidate_names", '')
                # Check if candidate_names is a list and convert to a string if it is
                if isinstance(candidate_names, list):
                    candidate_names_str = '|'.join(candidate_names)
                else:
                    candidate_names_str = candidate_names
                sheet.cell(row=next_row, column=i, value=candidate_names_str)
            
            # "GEO_method", "GEO_formatted_full_string", "GEO_decimal_lat", "GEO_decimal_long",
            # "GEO_city", "GEO_county", "GEO_state", "GEO_state_code", "GEO_country", "GEO_country_code", "GEO_continent"
            elif header.value in self.geo_headers:
                sheet.cell(row=next_row, column=i, value=GEO_record.get(header.value, ''))

            elif header.value in self.usage_headers:
                sheet.cell(row=next_row, column=i, value=usage_report.get(header.value, ''))

            elif header.value == "LLM":
                sheet.cell(row=next_row, column=i, value=MODEL_NAME_FORMATTED)

        # save the workbook
        wb.save(path_transcription)
    

    def has_API_key(self, val):
        return isinstance(val, str) and bool(val.strip())
        # if val != '':
        #     return True
        # else:
        #     return False
        

    def get_google_credentials(self): # Also used for google drive
        if self.is_hf or self.skip_API_keys:
            creds_json_str = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
            credentials = service_account.Credentials.from_service_account_info(json.loads(creds_json_str))
            return credentials
        else:
            with open(self.cfg_private['google']['GOOGLE_APPLICATION_CREDENTIALS'], 'r') as file:
                data = json.load(file)
                creds_json_str = json.dumps(data)
                credentials = service_account.Credentials.from_service_account_info(json.loads(creds_json_str))
                os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = creds_json_str
                return credentials
        

    def set_API_keys(self):
        if not self.skip_API_keys: 
            if self.is_hf:
                self.dir_home = os.path.dirname(os.path.dirname(__file__))
                self.path_cfg_private = None
                self.cfg_private = None

                k_openai = os.getenv('OPENAI_API_KEY')
                k_openai_azure = os.getenv('AZURE_API_VERSION')

                k_huggingface = None

                k_google_project_id = os.getenv('GOOGLE_PROJECT_ID')
                k_google_location = os.getenv('GOOGLE_LOCATION')
                k_google_application_credentials = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')

                k_hyperbolic = os.getenv('HYPERBOLIC_API_KEY')
                k_mistral = os.getenv('MISTRAL_API_KEY')
                k_here = os.getenv('HERE_API_KEY')
                k_opencage = os.getenv('open_cage_geocode')
            else:
                self.dir_home = os.path.dirname(os.path.dirname(__file__))
                self.path_cfg_private = os.path.join(self.dir_home, 'PRIVATE_DATA.yaml')
                if not os.path.exists(self.path_cfg_private):
                    og_pvt = self.path_cfg_private
                    self.path_cfg_private = os.path.join(os.path.dirname(self.dir_home), 'PRIVATE_DATA.yaml')
                    if not os.path.exists(self.path_cfg_private):
                        raise FileNotFoundError(f"Cannot find PRIVATE_DATA.yaml in either of these locations: {og_pvt} OR {self.path_cfg_private}")
                
                self.cfg_private = get_cfg_from_full_path(self.path_cfg_private)

                k_openai = self.cfg_private['openai']['OPENAI_API_KEY']
                k_openai_azure = self.cfg_private['openai_azure']['OPENAI_API_KEY_AZURE']

                k_huggingface = self.cfg_private['huggingface']['hf_token']
                os.environ["HUGGING_FACE_KEY"] = k_huggingface

                k_google_project_id = self.cfg_private['google']['GOOGLE_PROJECT_ID']
                k_google_location = self.cfg_private['google']['GOOGLE_LOCATION']
                k_google_application_credentials = self.cfg_private['google']['GOOGLE_APPLICATION_CREDENTIALS']
                
                k_hyperbolic = self.cfg_private['hyperbolic']['HYPERBOLIC_API_KEY']
                k_mistral = self.cfg_private['mistral']['MISTRAL_API_KEY']
                k_here = self.cfg_private['here']['API_KEY']
                k_opencage = self.cfg_private['open_cage_geocode']['API_KEY']
        else: # For VVGO
            self.path_cfg_private = None
            self.cfg_private = None

            k_openai = None
            k_openai_azure = None

            k_huggingface = None

            k_google_project_id = os.getenv('GOOGLE_PROJECT_ID')
            k_google_location = os.getenv('GOOGLE_LOCATION')
            k_google_application_credentials = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')

            k_hyperbolic = None
            k_mistral = None
            k_here = None
            k_opencage = None


        self.has_key_openai = self.has_API_key(k_openai)
        self.has_key_azure_openai = self.has_API_key(k_openai_azure)
        self.llm = None
        
        self.has_key_huggingface = self.has_API_key(k_huggingface)

        self.has_key_google_project_id = self.has_API_key(k_google_project_id)
        self.has_key_google_location = self.has_API_key(k_google_location)
        self.has_key_google_application_credentials = self.has_API_key(k_google_application_credentials)

        self.has_key_hyperbolic = self.has_API_key(k_hyperbolic)
        self.has_key_mistral = self.has_API_key(k_mistral)
        self.has_key_here = self.has_API_key(k_here)
        self.has_key_open_cage_geocode = self.has_API_key(k_opencage)

        

        ### Google - OCR, Palm2, Gemini
        if self.has_key_google_application_credentials and self.has_key_google_project_id and self.has_key_google_location:
            if self.is_hf or self.skip_API_keys:
                vertexai.init(project=os.getenv('GOOGLE_PROJECT_ID'), location=os.getenv('GOOGLE_LOCATION'), credentials=self.get_google_credentials())
            else:
                vertexai.init(project=k_google_project_id, location=k_google_location, credentials=self.get_google_credentials())
                os.environ['GOOGLE_API_KEY'] = self.cfg_private['google']['GOOGLE_PALM_API']


        ### OpenAI
        if self.has_key_openai:
            if self.is_hf:
                openai.api_key = os.getenv('OPENAI_API_KEY')
            else:
                openai.api_key = self.cfg_private['openai']['OPENAI_API_KEY']
                os.environ["OPENAI_API_KEY"] = self.cfg_private['openai']['OPENAI_API_KEY']

        if self.has_key_huggingface:
            if self.is_hf:
                pass
            else:
                os.environ["HUGGING_FACE_KEY"] = self.cfg_private['huggingface']['hf_token']

        ### OpenAI - Azure
        if self.has_key_azure_openai:
            if self.is_hf:
                # Initialize the Azure OpenAI client
                self.llm = AzureChatOpenAI(
                    deployment_name = 'gpt-35-turbo',#'gpt-35-turbo',
                    openai_api_version = os.getenv('AZURE_API_VERSION'),
                    openai_api_key = os.getenv('AZURE_API_KEY'),
                    azure_endpoint = os.getenv('AZURE_API_BASE'),
                    openai_organization = os.getenv('AZURE_ORGANIZATION'),
                )
                
            else:
                # Initialize the Azure OpenAI client
                self.llm = AzureChatOpenAI(
                    deployment_name = 'gpt-35-turbo',#'gpt-35-turbo',
                    openai_api_version = self.cfg_private['openai_azure']['OPENAI_API_VERSION'],
                    openai_api_key = self.cfg_private['openai_azure']['OPENAI_API_KEY_AZURE'],
                    azure_endpoint = self.cfg_private['openai_azure']['OPENAI_API_BASE'],
                    openai_organization = self.cfg_private['openai_azure']['OPENAI_ORGANIZATION'],
                )
                

        ### Mistral
        if self.has_key_mistral:
            if self.is_hf:
                pass # Already set
            else:
                os.environ['MISTRAL_API_KEY'] = self.cfg_private['mistral']['MISTRAL_API_KEY']
        

        ### hyperbolic
        if self.has_key_hyperbolic:
            if self.is_hf:
                pass # Already set
            else:
                os.environ['HYPERBOLIC_API_KEY'] = self.cfg_private['hyperbolic']['HYPERBOLIC_API_KEY']


        ### HERE
        if self.has_key_here:
            if self.is_hf:
                pass # Already set
            else:
                os.environ['HERE_APP_ID'] = self.cfg_private['here']['APP_ID']
                os.environ['HERE_API_KEY'] = self.cfg_private['here']['API_KEY']


        ### HERE
        if self.has_key_open_cage_geocode:
            if self.is_hf:
                pass # Already set
            else:
                os.environ['OPENCAGE_API_KEY'] = self.cfg_private['open_cage_geocode']['API_KEY']
                

        
    def clean_catalog_number(self, data, filename_without_extension):
        #Cleans up the catalog number in data if it's a dict
        
        def modify_catalog_key(catalog_key, filename_without_extension, data):
            # Helper function to apply modifications on catalog number
            if catalog_key not in data:
                new_data = {catalog_key: None}
                data = {**new_data, **data}

            if self.prefix_removal:
                filename_without_extension = filename_without_extension.replace(self.prefix_removal, "")
            if self.suffix_removal:
                filename_without_extension = filename_without_extension.replace(self.suffix_removal, "")
            if self.catalog_numerical_only:
                filename_without_extension = self.remove_non_numbers(data[catalog_key])
            data[catalog_key] = filename_without_extension
            return data
        
        if isinstance(data, dict):
            if self.headers_used == 'HEADERS_v1_n22':
                return modify_catalog_key("Catalog Number", filename_without_extension, data)
            elif self.headers_used in ['HEADERS_v2_n26', 'CUSTOM']:
                return modify_catalog_key("filename", filename_without_extension, data)
            else:
                raise ValueError("Invalid headers used.")
        else:
            raise TypeError("Data is not of type dict.")
        

    def write_json_to_file(self, filepath, data):
        '''Writes dictionary data to a JSON file.'''
        with open(filepath, 'w') as txt_file:
            if isinstance(data, dict):
                data = json.dumps(data, indent=4, sort_keys=False)
            txt_file.write(data)


    # def create_null_json(self):
    #     return {}
    

    def remove_non_numbers(self, s):
        return ''.join([char for char in s if char.isdigit()])
    

    def create_null_row(self, filename_without_extension, path_to_crop, path_to_content, path_to_helper):
        json_dict = {header: '' for header in self.headers} 
        for header, value in json_dict.items():
            if header == "path_to_crop":
                json_dict[header] = path_to_crop
            elif header == "path_to_original":
                fname = os.path.basename(path_to_crop)
                base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(path_to_crop))))
                path_to_original = os.path.join(base, 'Original_Images', fname)
                json_dict[header] = path_to_original
            elif header == "path_to_content":
                json_dict[header] = path_to_content
            elif header == "path_to_helper":
                json_dict[header] = path_to_helper
            elif header == "filename":
                json_dict[header] = filename_without_extension

            # "WFO_exact_match","WFO_exact_match_name","WFO_best_match","WFO_candidate_names","WFO_placement"
            elif header == "WFO_exact_match":
                json_dict[header] =''
            elif header == "WFO_exact_match_name":
                json_dict[header] = ''
            elif header == "WFO_best_match":
                json_dict[header] = ''
            elif header == "WFO_candidate_names":
                json_dict[header] = ''
            elif header == "WFO_placement":
                json_dict[header] = ''
        return json_dict
    

    ##################################################################################################################################
    ##################################################     OCR      ##################################################################
    ##################################################################################################################################
    def perform_OCR_and_save_results(self, image_index, json_report, jpg_file_path_OCR_helper, txt_file_path_OCR, txt_file_path_OCR_bounds, path_to_crop, OCR_Engine):
        
        self.logger.info(f'Working on {image_index + 1}/{len(self.img_paths)} --- Starting OCR')
        # self.OCR - None
        self.OCR_cost = 0.0
        self.OCR_tokens_in = 0
        self.OCR_tokens_out = 0
        self.OCR_cost_in = 0
        self.OCR_cost_out = 0

        ### Process_image() runs the OCR for text, handwriting, trOCR AND creates the overlay image
        OCR_Engine.process_image(self.do_create_OCR_helper_image, path_to_crop, self.logger)
        self.OCR = OCR_Engine.OCR

        self.OCR_cost = OCR_Engine.cost
        self.OCR_tokens_in = OCR_Engine.tokens_in
        self.OCR_tokens_out = OCR_Engine.tokens_out

        self.logger.info(f"Complete OCR text for LLM prompt:\n\n{self.OCR}\n\n")

        self.write_json_to_file(txt_file_path_OCR, OCR_Engine.OCR_JSON_to_file)
        
        self.logger.info(f'Working on {image_index + 1}/{len(self.img_paths)} --- Finished OCR')

        if len(self.OCR) > 0:
            OCR_Engine.overlay_image.save(jpg_file_path_OCR_helper)

            OCR_bounds = {}
            if OCR_Engine.hand_text_to_box_mapping is not None:
                OCR_bounds['OCR_bounds_handwritten'] = OCR_Engine.hand_text_to_box_mapping

            if OCR_Engine.normal_text_to_box_mapping is not None:
                OCR_bounds['OCR_bounds_printed'] = OCR_Engine.normal_text_to_box_mapping

            if OCR_Engine.trOCR_text_to_box_mapping is not None:
                OCR_bounds['OCR_bounds_trOCR'] = OCR_Engine.trOCR_text_to_box_mapping

            self.write_json_to_file(txt_file_path_OCR_bounds, OCR_bounds)
            self.logger.info(f'Working on {image_index + 1}/{len(self.img_paths)} --- Saved OCR Overlay Image')
        else:
            pass ########################################################################################################################### fix logic for no OCR

    ##################################################################################################################################
    #######################################################  LLM Switchboard  ########################################################
    ##################################################################################################################################
    # def send_to_LLM(self, is_azure, progress_report, json_report, model_name):
    #     self.n_failed_LLM_calls = 0
    #     self.n_failed_OCR = 0

    #     final_JSON_response = None
    #     final_WFO_record = None
    #     final_GEO_record = None

    #     self.initialize_token_counters()
    #     self.update_progress_report_initial(progress_report)

    #     MODEL_NAME_FORMATTED = ModelMaps.get_API_name(model_name)
    #     name_parts = model_name.split("_")
        
    #     self.setup_JSON_dict_structure()

    #     Copy_Prompt = PromptCatalog()
    #     Copy_Prompt.copy_prompt_template_to_new_dir(self.Dirs.transcription_prompt, self.path_custom_prompts)
        
    #     if json_report:
    #         json_report.set_text(text_main=f'Loading {MODEL_NAME_FORMATTED}')
    #         json_report.set_JSON({}, {}, {})
    #     llm_model = self.initialize_llm_model(self.cfg, self.logger, MODEL_NAME_FORMATTED, self.JSON_dict_structure, name_parts, is_azure, self.llm, self.config_vals_for_permutation)

    #     OCR_Engine = OCREngine(self.logger, json_report, self.dir_home, self.is_hf, self.cfg, self.trOCR_model_version, self.trOCR_model, self.trOCR_processor, self.device)  
        

    #     for i, path_to_crop in enumerate(self.img_paths):
    #         self.update_progress_report_batch(progress_report, i)

    #         if self.should_skip_specimen(path_to_crop):
    #             self.log_skipping_specimen(path_to_crop)
    #             continue

    #         paths = self.generate_paths(path_to_crop, i)
    #         self.path_to_crop = path_to_crop

    #         filename_without_extension, txt_file_path, txt_file_path_OCR, txt_file_path_OCR_bounds, jpg_file_path_OCR_helper, json_file_path_wiki, txt_file_path_ind_prompt = paths
    #         if json_report:
    #             json_report.set_text(text_main='Starting OCR')
    #         self.perform_OCR_and_save_results(i, json_report, jpg_file_path_OCR_helper, txt_file_path_OCR, txt_file_path_OCR_bounds, self.path_to_crop, OCR_Engine)
    #         self.OCR_cost += OCR_Engine.cost
    #         self.OCR_tokens_in += OCR_Engine.tokens_in
    #         self.OCR_tokens_out += OCR_Engine.tokens_out
    #         self.OCR_cost_in += OCR_Engine.cost_in
    #         self.OCR_cost_out += OCR_Engine.cost_out
    #         self.ocr_method = str(OCR_Engine.ocr_method)

    #         if json_report:
    #             json_report.set_text(text_main='Finished OCR')

    #         if not self.OCR:
    #             self.n_failed_OCR += 1
    #             response_candidate = None
    #             nt_in = 0
    #             nt_out = 0
    #         else:
    #             ### Format prompt
    #             prompt = self.setup_prompt()
    #             # prompt = remove_colons_and_double_apostrophes(prompt) # This is moved to utils_VV since it broke the json structure.

    #             ### Send prompt to chosen LLM
    #             self.logger.info(f'Waiting for {model_name} API call --- Using {MODEL_NAME_FORMATTED}')

    #             if 'PALM2' in name_parts:
    #                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_GooglePalm2(prompt, json_report, paths)
                
    #             elif 'GEMINI' in name_parts:
    #                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_GoogleGemini(prompt, json_report, paths)
                
    #             elif 'MISTRAL' in name_parts and ('LOCAL' not in name_parts):
    #                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_MistralAI(prompt, json_report, paths)

    #             elif 'Hyperbolic' in name_parts:
    #                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_Hyperbolic(prompt, json_report, paths)
                
                
    #             elif 'LOCAL' in name_parts: 
    #                 if 'MISTRAL' in name_parts or 'MIXTRAL' in name_parts:
    #                     if 'CPU' in name_parts:     
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_local_cpu_MistralAI(prompt, json_report, paths) 
    #                     else:
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_local_MistralAI(prompt, json_report, paths) 
                
    #             elif "/" in ''.join(name_parts):
    #                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_local_custom_fine_tune(self.OCR, json_report, paths)  ###

    #             else:
    #                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_OpenAI(prompt, json_report, paths)

    #         self.n_failed_LLM_calls += 1 if response_candidate is None else 0
                
    #         ### Estimate n tokens returned
    #         self.logger.info(f'Prompt tokens IN --- {nt_in}')
    #         self.logger.info(f'Prompt tokens OUT --- {nt_out}')
                
    #         self.update_token_counters(nt_in, nt_out)

    #         final_JSON_response, final_WFO_record, final_GEO_record = self.update_final_response(response_candidate, WFO_record, GEO_record, usage_report, MODEL_NAME_FORMATTED, paths, path_to_crop, nt_in, nt_out)

    #         self.logger.info(f'Finished LLM call')
            
    #         if json_report:
    #             json_report.set_JSON(final_JSON_response, final_WFO_record, final_GEO_record)

    #     del OCR_Engine
    #     torch.cuda.empty_cache()
    #     gc.collect()

    #     self.update_progress_report_final(progress_report)
    #     final_JSON_response = self.parse_final_json_response(final_JSON_response)
        
    #     return final_JSON_response, final_WFO_record, final_GEO_record, self.total_tokens_in, self.total_tokens_out, self.OCR_cost, self.OCR_tokens_in, self.OCR_tokens_out, self.OCR_cost_in, self.OCR_cost_out, self.ocr_method

    

    ##################################################################################################################################
    #######################################################  LLM Switchboard  ########################################################
    ################################ This version is parallelized and WILL NOT work with local LLMs  #################################
    ##################################################################################################################################
    # def send_to_LLM(self, is_azure, progress_report, json_report, model_name):
    #     self.n_failed_LLM_calls = 0
    #     self.n_failed_OCR = 0

    #     final_JSON_response = None
    #     final_WFO_record = None
    #     final_GEO_record = None

    #     self.initialize_token_counters()
    #     self.update_progress_report_initial(progress_report)

    #     MODEL_NAME_FORMATTED = ModelMaps.get_API_name(model_name)
    #     name_parts = model_name.split("_")
        
    #     self.setup_JSON_dict_structure()

    #     Copy_Prompt = PromptCatalog()
    #     Copy_Prompt.copy_prompt_template_to_new_dir(self.Dirs.transcription_prompt, self.path_custom_prompts)
        
    #     if json_report:
    #         json_report.set_text(text_main=f'Loading {MODEL_NAME_FORMATTED}')
    #         json_report.set_JSON({}, {}, {})
                
    #     llm_model = self.initialize_llm_model(self.cfg, self.logger, MODEL_NAME_FORMATTED, 
    #                                         self.JSON_dict_structure, name_parts, is_azure, 
    #                                         self.llm, self.config_vals_for_permutation)

    #     # Create an OCR Engine that will be used by the main thread
    #     main_OCR_Engine = OCREngine(self.logger, json_report, self.dir_home, self.is_hf, 
    #                                 self.cfg, self.trOCR_model_version, self.trOCR_model, 
    #                                 self.trOCR_processor, self.device)
        
    #     # Filter out specimens that should be skipped
    #     valid_img_paths = []
    #     for i, path_to_crop in enumerate(self.img_paths):
    #         if not self.should_skip_specimen(path_to_crop):
    #             valid_img_paths.append((i, path_to_crop))
    #         else:
    #             self.log_skipping_specimen(path_to_crop)
        
    #     # Create job queue and result list with locks
    #     job_queue = queue.Queue()
    #     results = {}
    #     excel_rows = []  # To store data for Excel file
    #     results_lock = threading.Lock()
    #     ocr_stats_lock = threading.Lock()  # For updating OCR related statistics
    #     excel_rows_lock = threading.Lock()  # For updating excel rows
        
    #     # Add all jobs to the queue
    #     for i, path_to_crop in valid_img_paths:
    #         job_queue.put((i, path_to_crop))
        
    #     # Split the save_json_and_xlsx function into save_json and prepare_excel_row
    #     def save_json(response, filename_without_extension, txt_file_path):
    #         """Save JSON to file, but don't update Excel"""
    #         if response is None:
    #             response = self.JSON_dict_structure.copy()
    #             # Insert 'filename' as the first key
    #             response = {'filename': filename_without_extension, **{k: v for k, v in response.items() if k != 'filename'}}
    #         else:
    #             response = self.clean_catalog_number(response, filename_without_extension)
            
    #         self.write_json_to_file(txt_file_path, response)
    #         return response
        
    #     def prepare_excel_row(response, WFO_record, GEO_record, usage_report, 
    #                         MODEL_NAME_FORMATTED, filename_without_extension, path_to_crop, 
    #                         txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out):
    #         """Prepare a row for Excel without writing to file"""
    #         # Create a dictionary to store the Excel row data
    #         row_data = {
    #             # Basic fields
    #             'filename': filename_without_extension,
    #             'path_to_crop': path_to_crop,
    #             'path_to_content': txt_file_path,
    #             'path_to_helper': jpg_file_path_OCR_helper,
    #             'tokens_in': nt_in,
    #             'tokens_out': nt_out,
    #             'prompt': os.path.basename(self.path_custom_prompts),
    #             'run_name': self.Dirs.run_name,
    #             'LM2_collage': self.cfg['leafmachine']['use_RGB_label_images'],
    #             'LLM': MODEL_NAME_FORMATTED
    #         }
            
    #         # OCR related fields
    #         ocr_method = self.cfg['leafmachine']['project']['OCR_option']
    #         if isinstance(ocr_method, list):
    #             ocr_method = '|'.join(map(str, ocr_method))
            
    #         row_data['OCR_method'] = ocr_method
    #         row_data['OCR_double'] = self.cfg['leafmachine']['project']['double_OCR']
    #         row_data['OCR_trOCR'] = self.cfg['leafmachine']['project']['do_use_trOCR']
            
    #         # Path to original
    #         if self.cfg['leafmachine']['use_RGB_label_images'] in [1, 2]:
    #             fname = os.path.basename(path_to_crop)
    #             base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(path_to_crop))))
    #             path_to_original = os.path.join(base, 'Original_Images', fname)
    #         else:
    #             fname = os.path.basename(path_to_crop)
    #             base = os.path.dirname(os.path.dirname(path_to_crop))
    #             path_to_original = os.path.join(base, 'Original_Images', fname)
            
    #         row_data['path_to_original'] = path_to_original
            
    #         # Process response data
    #         if response is not None:
    #             if isinstance(response, str):
    #                 try:
    #                     response = json.loads(response)
    #                 except json.JSONDecodeError:
    #                     self.logger.error(f"Failed to parse response: {response}")
    #                     response = {}
                
    #             # Add response data to row_data
    #             for key, value in response.items():
    #                 if key not in self.catalog_name_options:
    #                     if isinstance(value, dict):
    #                         # If it's a dictionary, extract the 'value' field
    #                         row_data[key] = value.get('value', '')
    #                     else:
    #                         # If it's not a dictionary, use it directly
    #                         row_data[key] = value
                
    #         # Add WFO data
    #         if WFO_record:
    #             for header in self.wfo_headers_no_lists:
    #                 if header in WFO_record:
    #                     row_data[header] = WFO_record[header]
                
    #             # Handle WFO_candidate_names specially
    #             candidate_names = WFO_record.get("WFO_candidate_names", '')
    #             if isinstance(candidate_names, list):
    #                 row_data["WFO_candidate_names"] = '|'.join(candidate_names)
    #             else:
    #                 row_data["WFO_candidate_names"] = candidate_names
            
    #         # Add GEO data
    #         if GEO_record:
    #             for header in self.geo_headers:
    #                 if header in GEO_record:
    #                     row_data[header] = GEO_record[header]
            
    #         # Add usage data
    #         if usage_report:
    #             for header in self.usage_headers:
    #                 if header in usage_report:
    #                     row_data[header] = usage_report[header]
            
    #         return row_data
        
    #     # Worker function
    #     def worker():
    #         # Each worker gets its own OCR Engine
    #         worker_OCR_Engine = None  # Lazy initialization
    #         worker_id = threading.current_thread().name
    #         json_report = None
            
    #         while True:
    #             i = None
    #             path_to_crop = None
                
    #             try:
    #                 # Get a job from the queue, non-blocking
    #                 try:
    #                     i, path_to_crop = job_queue.get(block=False)
    #                 except queue.Empty:
    #                     # No more jobs
    #                     break
                    
    #                 self.logger.info(f'[{worker_id}] Working on {i+1}/{len(valid_img_paths)} --- {os.path.basename(path_to_crop)}')
                    
    #                 # Initialize worker's OCR Engine if needed
    #                 if worker_OCR_Engine is None:
    #                     try:
    #                         worker_OCR_Engine = OCREngine(self.logger, json_report, self.dir_home, self.is_hf, 
    #                                                     self.cfg, self.trOCR_model_version, self.trOCR_model, 
    #                                                     self.trOCR_processor, self.device)
    #                         self.logger.info(f'[{worker_id}] Successfully initialized OCR Engine')
    #                     except Exception as e:
    #                         self.logger.error(f'[{worker_id}] Error initializing OCR Engine: {str(e)}')
    #                         self.logger.error(traceback.format_exc())
    #                         # Use main OCR Engine as fallback
    #                         worker_OCR_Engine = main_OCR_Engine
    #                         self.logger.info(f'[{worker_id}] Using main OCR Engine as fallback')
                    
    #                 # Generate paths
    #                 paths = self.generate_paths(path_to_crop, i)
    #                 filename_without_extension, txt_file_path, txt_file_path_OCR, txt_file_path_OCR_bounds, jpg_file_path_OCR_helper, json_file_path_wiki, txt_file_path_ind_prompt = paths
                    
    #                 # Verify the image file exists
    #                 if not os.path.isfile(path_to_crop):
    #                     self.logger.error(f'[{worker_id}] Image file not found: {path_to_crop}')
                        
    #                     # Save JSON for failed image processing
    #                     null_response = save_json(None, filename_without_extension, txt_file_path)
                        
    #                     # Prepare Excel row
    #                     row_data = prepare_excel_row(
    #                         null_response, None, None, None, MODEL_NAME_FORMATTED,
    #                         filename_without_extension, path_to_crop, txt_file_path, 
    #                         jpg_file_path_OCR_helper, 0, 0
    #                     )
                        
    #                     with excel_rows_lock:
    #                         excel_rows.append(row_data)
                        
    #                     with results_lock:
    #                         results[i] = {
    #                             'path_to_crop': path_to_crop,
    #                             'ocr_failed': True,
    #                             'llm_failed': False,
    #                             'nt_in': 0,
    #                             'nt_out': 0,
    #                             'response': null_response,
    #                             'wfo_record': None,
    #                             'geo_record': None,
    #                             'usage_report': None,
    #                             'error': 'Image file not found'
    #                         }
    #                         self.n_failed_OCR += 1
                        
    #                     job_queue.task_done()
    #                     continue
                    
    #                 # Perform OCR with detailed error handling
    #                 if json_report:
    #                     json_report.set_text(text_main=f'Starting OCR for {os.path.basename(path_to_crop)}')
                    
    #                 ocr_success = False
    #                 try:
    #                     self.logger.info(f'[{worker_id}] Starting OCR for {os.path.basename(path_to_crop)}')
    #                     self.perform_OCR_and_save_results(i, json_report, jpg_file_path_OCR_helper, 
    #                                                     txt_file_path_OCR, txt_file_path_OCR_bounds, 
    #                                                     path_to_crop, worker_OCR_Engine)
                        
    #                     # Capture OCR stats
    #                     ocr_cost = worker_OCR_Engine.cost
    #                     ocr_tokens_in = worker_OCR_Engine.tokens_in
    #                     ocr_tokens_out = worker_OCR_Engine.tokens_out
    #                     ocr_cost_in = worker_OCR_Engine.cost_in
    #                     ocr_cost_out = worker_OCR_Engine.cost_out
    #                     ocr_method = str(worker_OCR_Engine.ocr_method)
    #                     ocr_success = bool(self.OCR)
                        
    #                     # Update OCR stats atomically
    #                     with ocr_stats_lock:
    #                         self.OCR_cost += ocr_cost
    #                         self.OCR_tokens_in += ocr_tokens_in
    #                         self.OCR_tokens_out += ocr_tokens_out
    #                         self.OCR_cost_in += ocr_cost_in
    #                         self.OCR_cost_out += ocr_cost_out
    #                         self.ocr_method = ocr_method
    #                         if not ocr_success:
    #                             self.n_failed_OCR += 1
                        
    #                     self.logger.info(f'[{worker_id}] OCR {"succeeded" if ocr_success else "failed"} for {os.path.basename(path_to_crop)}')
    #                 except Exception as e:
    #                     self.logger.error(f'[{worker_id}] OCR error for {os.path.basename(path_to_crop)}: {str(e)}')
    #                     self.logger.error(traceback.format_exc())
    #                     ocr_success = False
    #                     with ocr_stats_lock:
    #                         self.n_failed_OCR += 1
                    
    #                 if json_report:
    #                     json_report.set_text(text_main=f'Finished OCR for {os.path.basename(path_to_crop)}')
                    
    #                 # If OCR failed, save JSON for failed OCR and continue
    #                 if not ocr_success:
    #                     # Save JSON for failed OCR
    #                     null_response = save_json(None, filename_without_extension, txt_file_path)
                        
    #                     # Prepare Excel row
    #                     row_data = prepare_excel_row(
    #                         null_response, None, None, None, MODEL_NAME_FORMATTED,
    #                         filename_without_extension, path_to_crop, txt_file_path, 
    #                         jpg_file_path_OCR_helper, 0, 0
    #                     )
                        
    #                     with excel_rows_lock:
    #                         excel_rows.append(row_data)
                        
    #                     with results_lock:
    #                         results[i] = {
    #                             'path_to_crop': path_to_crop,
    #                             'ocr_failed': True,
    #                             'llm_failed': False,
    #                             'nt_in': 0,
    #                             'nt_out': 0,
    #                             'response': null_response,
    #                             'wfo_record': None,
    #                             'geo_record': None,
    #                             'usage_report': None,
    #                             'error': 'OCR failed'
    #                         }
    #                     job_queue.task_done()
    #                     continue
                    
    #                 # Setup prompt for LLM
    #                 try:
    #                     prompt = self.setup_prompt()
    #                     self.logger.info(f'[{worker_id}] Successfully created prompt for {os.path.basename(path_to_crop)}')
    #                 except Exception as e:
    #                     self.logger.error(f'[{worker_id}] Error creating prompt for {os.path.basename(path_to_crop)}: {str(e)}')
    #                     self.logger.error(traceback.format_exc())
                        
    #                     # Save JSON for failed prompt creation
    #                     null_response = save_json(None, filename_without_extension, txt_file_path)
                        
    #                     # Prepare Excel row
    #                     row_data = prepare_excel_row(
    #                         null_response, None, None, None, MODEL_NAME_FORMATTED,
    #                         filename_without_extension, path_to_crop, txt_file_path, 
    #                         jpg_file_path_OCR_helper, 0, 0
    #                     )
                        
    #                     with excel_rows_lock:
    #                         excel_rows.append(row_data)
                        
    #                     with results_lock:
    #                         results[i] = {
    #                             'path_to_crop': path_to_crop,
    #                             'ocr_failed': False,
    #                             'llm_failed': True,
    #                             'nt_in': 0,
    #                             'nt_out': 0,
    #                             'response': null_response,
    #                             'wfo_record': None,
    #                             'geo_record': None,
    #                             'usage_report': None,
    #                             'error': f'Prompt creation error: {str(e)}'
    #                         }
    #                         self.n_failed_LLM_calls += 1
    #                     job_queue.task_done()
    #                     continue
                    
    #                 # LLM call with detailed error handling
    #                 self.logger.info(f'[{worker_id}] Waiting for {model_name} API call for {os.path.basename(path_to_crop)} --- Using {MODEL_NAME_FORMATTED}')
                    
    #                 response_candidate = None
    #                 nt_in = 0
    #                 nt_out = 0
    #                 WFO_record = None
    #                 GEO_record = None
    #                 usage_report = None
    #                 llm_error = None
                    
    #                 try:
    #                     # Call the appropriate LLM based on model name
    #                     if 'PALM2' in name_parts:
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_GooglePalm2(prompt, json_report, paths)
                        
    #                     elif 'GEMINI' in name_parts:
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_GoogleGemini(prompt, json_report, paths)
                        
    #                     elif 'MISTRAL' in name_parts and ('LOCAL' not in name_parts):
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_MistralAI(prompt, json_report, paths)
                        
    #                     elif 'Hyperbolic' in name_parts:
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_Hyperbolic(prompt, json_report, paths)
                        
    #                     elif 'LOCAL' in name_parts:
    #                         if 'MISTRAL' in name_parts or 'MIXTRAL' in name_parts:
    #                             if 'CPU' in name_parts:     
    #                                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_local_cpu_MistralAI(prompt, json_report, paths) 
    #                             else:
    #                                 response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_local_MistralAI(prompt, json_report, paths) 
                        
    #                     elif "/" in ''.join(name_parts):
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_local_custom_fine_tune(self.OCR, json_report, paths)
                        
    #                     else:
    #                         response_candidate, nt_in, nt_out, WFO_record, GEO_record, usage_report = llm_model.call_llm_api_OpenAI(prompt, json_report, paths)
                        
    #                     llm_failed = response_candidate is None
                        
    #                     if llm_failed:
    #                         with results_lock:
    #                             self.n_failed_LLM_calls += 1
    #                         self.logger.error(f'[{worker_id}] LLM call failed for {os.path.basename(path_to_crop)}')
    #                     else:
    #                         self.logger.info(f'[{worker_id}] LLM call succeeded for {os.path.basename(path_to_crop)}')
    #                         self.logger.info(f'[{worker_id}] Prompt tokens IN: {nt_in}, Prompt tokens OUT: {nt_out}')
                        
    #                 except Exception as e:
    #                     self.logger.error(f'[{worker_id}] Error in LLM call for {os.path.basename(path_to_crop)}: {str(e)}')
    #                     self.logger.error(traceback.format_exc())
    #                     llm_error = str(e)
    #                     llm_failed = True
    #                     with results_lock:
    #                         self.n_failed_LLM_calls += 1
                    
    #                 # Save the JSON response immediately, but prepare Excel row for later
    #                 try:
    #                     # Update token counters atomically
    #                     with results_lock:
    #                         self.update_token_counters(nt_in, nt_out)
                        
    #                     # Just save JSON (don't update Excel)
    #                     final_response = save_json(response_candidate, filename_without_extension, txt_file_path)
                        
    #                     # Prepare Excel row for later
    #                     row_data = prepare_excel_row(
    #                         final_response, WFO_record, GEO_record, usage_report, MODEL_NAME_FORMATTED,
    #                         filename_without_extension, path_to_crop, txt_file_path, 
    #                         jpg_file_path_OCR_helper, nt_in, nt_out
    #                     )
                        
    #                     with excel_rows_lock:
    #                         excel_rows.append(row_data)
                        
    #                     self.logger.info(f'[{worker_id}] Saved JSON for {os.path.basename(path_to_crop)}')
                        
    #                     # Store result
    #                     with results_lock:
    #                         results[i] = {
    #                             'path_to_crop': path_to_crop,
    #                             'paths': paths,
    #                             'ocr_failed': False,
    #                             'llm_failed': llm_failed,
    #                             'nt_in': nt_in,
    #                             'nt_out': nt_out,
    #                             'response': final_response,
    #                             'wfo_record': WFO_record,
    #                             'geo_record': GEO_record,
    #                             'usage_report': usage_report,
    #                             'error': llm_error
    #                         }
                        
    #                 except Exception as e:
    #                     self.logger.error(f'[{worker_id}] Error saving JSON for {os.path.basename(path_to_crop)}: {str(e)}')
    #                     self.logger.error(traceback.format_exc())
                        
    #                     # Try one more time to save a null response
    #                     try:
    #                         null_response = save_json(None, filename_without_extension, txt_file_path)
                            
    #                         # Prepare Excel row
    #                         row_data = prepare_excel_row(
    #                             null_response, None, None, None, MODEL_NAME_FORMATTED,
    #                             filename_without_extension, path_to_crop, txt_file_path, 
    #                             jpg_file_path_OCR_helper, nt_in, nt_out
    #                         )
                            
    #                         with excel_rows_lock:
    #                             excel_rows.append(row_data)
                            
    #                         with results_lock:
    #                             results[i] = {
    #                                 'path_to_crop': path_to_crop,
    #                                 'ocr_failed': False,
    #                                 'llm_failed': True,
    #                                 'nt_in': nt_in,
    #                                 'nt_out': nt_out,
    #                                 'response': null_response,
    #                                 'wfo_record': None,
    #                                 'geo_record': None,
    #                                 'usage_report': None,
    #                                 'error': f'JSON saving error: {str(e)}'
    #                             }
    #                     except Exception as save_error:
    #                         self.logger.error(f'[{worker_id}] Failed to save even null JSON for {os.path.basename(path_to_crop)}: {str(save_error)}')
                    
    #                 self.logger.info(f'[{worker_id}] Finished processing {os.path.basename(path_to_crop)}')
                    
    #                 # Mark job as done
    #                 job_queue.task_done()
                    
    #             except Exception as e:
    #                 error_msg = f"Unexpected error in worker: {str(e)}"
    #                 self.logger.error(error_msg)
    #                 self.logger.error(traceback.format_exc())
                    
    #                 # Try to mark as done to avoid hanging
    #                 if i is not None:
    #                     # Try to save a null response for this image
    #                     try:
    #                         path_to_crop = path_to_crop or "unknown"
    #                         paths = self.generate_paths(path_to_crop, i)
    #                         filename_without_extension, txt_file_path, _, _, jpg_file_path_OCR_helper, _, _ = paths
                            
    #                         null_response = save_json(None, filename_without_extension, txt_file_path)
                            
    #                         # Prepare Excel row
    #                         row_data = prepare_excel_row(
    #                             null_response, None, None, None, MODEL_NAME_FORMATTED,
    #                             filename_without_extension, path_to_crop, txt_file_path, 
    #                             jpg_file_path_OCR_helper, 0, 0
    #                         )
                            
    #                         with excel_rows_lock:
    #                             excel_rows.append(row_data)
                            
    #                         with results_lock:
    #                             results[i] = {
    #                                 'path_to_crop': path_to_crop,
    #                                 'ocr_failed': True,
    #                                 'llm_failed': True,
    #                                 'nt_in': 0,
    #                                 'nt_out': 0,
    #                                 'response': null_response,
    #                                 'wfo_record': None,
    #                                 'geo_record': None,
    #                                 'usage_report': None,
    #                                 'error': error_msg
    #                             }
    #                     except:
    #                         self.logger.error(f"Could not save null response for failed job {i}")
                        
    #                     try:
    #                         job_queue.task_done()
    #                     except:
    #                         pass
        
    #     # Create worker threads (16 workers)
    #     worker_threads = []
    #     num_workers = min(16, len(valid_img_paths))  # Don't create more threads than jobs
    #     self.logger.info(f"Starting {num_workers} worker threads")
        
    #     for w in range(num_workers):
    #         t = threading.Thread(target=worker, name=f"Worker-{w+1}")
    #         t.daemon = True  # Make thread terminate when main thread ends
    #         worker_threads.append(t)
    #         t.start()
        
    #     # Wait for all jobs to be completed
    #     try:
    #         job_queue.join()
    #         self.logger.info("All jobs completed")
    #     except Exception as e:
    #         self.logger.error(f"Error waiting for jobs to complete: {str(e)}")
        
    #     # Now that all workers are done, save the Excel file with all collected rows
    #     try:
    #         self.logger.info("Saving Excel file with all results")
            
    #         # Check if we have any rows to save
    #         if excel_rows:
    #             # Create or load the Excel file
    #             excel_path = self.path_transcription
    #             if os.path.exists(excel_path):
    #                 try:
    #                     # Try to open the existing Excel file
    #                     wb = openpyxl.load_workbook(excel_path)
    #                     sheet = wb.active
                        
    #                     # Get the headers from the first row
    #                     headers = [cell.value for cell in sheet[1]]
                        
    #                     # Start from the next empty row
    #                     next_row = sheet.max_row + 1
                        
    #                     # For each prepared row, add it to the Excel file
    #                     for row_data in excel_rows:
    #                         for i, header in enumerate(headers, start=1):
    #                             if header in row_data:
    #                                 sheet.cell(row=next_row, column=i, value=row_data[header])
    #                         next_row += 1
                        
    #                     # Save the workbook
    #                     wb.save(excel_path)
    #                     self.logger.info(f"Successfully updated Excel file with {len(excel_rows)} new rows")
                    
    #                 except Exception as e:
    #                     self.logger.error(f"Error updating existing Excel file: {str(e)}")
    #                     self.logger.error(traceback.format_exc())
                        
    #                     # Try creating a new Excel file as fallback
    #                     try:
    #                         # Convert excel_rows to a pandas DataFrame
    #                         df = pd.DataFrame(excel_rows)
    #                         # Save to Excel
    #                         df.to_excel(excel_path, index=False)
    #                         self.logger.info(f"Created new Excel file with {len(excel_rows)} rows as fallback")
    #                     except Exception as e2:
    #                         self.logger.error(f"Failed to create fallback Excel file: {str(e2)}")
    #             else:
    #                 # File doesn't exist, create a new one
    #                 df = pd.DataFrame(excel_rows)
    #                 df.to_excel(excel_path, index=False)
    #                 self.logger.info(f"Created new Excel file with {len(excel_rows)} rows")
    #         else:
    #             self.logger.warning("No Excel rows to save")
        
    #     except Exception as e:
    #         self.logger.error(f"Error saving Excel file: {str(e)}")
    #         self.logger.error(traceback.format_exc())
        
    #     # Process results to update the progress report and find the final result
    #     for i, _ in valid_img_paths:
    #         if i in results:
    #             # Update progress report
    #             self.update_progress_report_batch(progress_report, i)
                
    #             # Get a result for final return value
    #             result = results[i]
    #             if not result['ocr_failed'] and not result['llm_failed']:
    #                 final_JSON_response = result['response']
    #                 final_WFO_record = result['wfo_record']
    #                 final_GEO_record = result['geo_record']
                    
    #                 if json_report:
    #                     json_report.set_JSON(final_JSON_response, final_WFO_record, final_GEO_record)
        
    #     # Clean up
    #     del main_OCR_Engine
    #     torch.cuda.empty_cache()
    #     gc.collect()
        
    #     self.update_progress_report_final(progress_report)
        
    #     try:
    #         if final_JSON_response:
    #             final_JSON_response = self.parse_final_json_response(final_JSON_response)
    #     except Exception as e:
    #         self.logger.error(f'Error parsing final JSON response: {str(e)}')
    #         self.logger.error(traceback.format_exc())
        
    #     return final_JSON_response, final_WFO_record, final_GEO_record, self.total_tokens_in, self.total_tokens_out, self.OCR_cost, self.OCR_tokens_in, self.OCR_tokens_out, self.OCR_cost_in, self.OCR_cost_out, self.ocr_method
    # --- Replace the existing send_to_LLM method inside the VoucherVision class ---

    def send_to_LLM(self, is_azure, progress_report, json_report, model_name):
        self.n_failed_LLM_calls = 0
        self.n_failed_OCR = 0
        self.initialize_token_counters()
        self.update_progress_report_initial(progress_report)

        MODEL_NAME_FORMATTED = ModelMaps.get_API_name(model_name)
        name_parts = model_name.split("_")

        self.setup_JSON_dict_structure()
        Copy_Prompt = PromptCatalog()
        Copy_Prompt.copy_prompt_template_to_new_dir(self.Dirs.transcription_prompt, self.path_custom_prompts)

        if json_report:
            json_report.set_text(text_main=f'Loading {MODEL_NAME_FORMATTED}')
            json_report.set_JSON({}, {}, {})

        # Initialize the LLM model once
        llm_model = self.initialize_llm_model(self.cfg, self.logger, MODEL_NAME_FORMATTED,
                                            self.JSON_dict_structure, name_parts, is_azure,
                                            self.llm, self.config_vals_for_permutation)

        # Filter out specimens that should be skipped
        valid_img_paths = []
        for i, path_to_crop in enumerate(self.img_paths):
            if not self.should_skip_specimen(path_to_crop):
                valid_img_paths.append((i, path_to_crop))
            else:
                self.log_skipping_specimen(path_to_crop)

        # --- Prepare Jobs for Parallel Processing ---
        jobs = []
        total_images = len(valid_img_paths)
        for i, path_to_crop in valid_img_paths:
            # Each thread gets its own fresh OCR Engine instance to prevent state conflicts
            ocr_engine_for_thread = OCREngine(self.logger, None, self.dir_home, self.is_hf,
                                            self.cfg, self.trOCR_model_version, self.trOCR_model,
                                            self.trOCR_processor, self.device)
            job_args = (
                i, path_to_crop, total_images, llm_model, ocr_engine_for_thread,
                self, MODEL_NAME_FORMATTED, name_parts # 'self' is passed for access to stateless helpers
            )
            jobs.append(job_args)

        # --- Execute Jobs in Parallel ---
        all_results = [None] * len(self.img_paths)
        num_workers = min(16, len(valid_img_paths))
        self.logger.info(f"Starting {num_workers} worker threads for {len(jobs)} images.")

        with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
            # Use a progress bar
            with tqdm(total=len(jobs), desc="Processing Images", unit="image") as pbar:
                # map jobs to the worker function
                future_to_job = {executor.submit(process_single_image_worker, job): job for job in jobs}

                for future in concurrent.futures.as_completed(future_to_job):
                    try:
                        result_package = future.result()
                        # Place the result in the correct position using its index
                        all_results[result_package['index']] = result_package
                    except Exception as e:
                        self.logger.error(f"A worker failed unexpectedly: {e}")
                        self.logger.error(traceback.format_exc())
                    finally:
                        pbar.update(1)

        # --- Process and Aggregate Results Serially and Safely ---
        final_JSON_response, final_WFO_record, final_GEO_record = None, None, None
        collected_ocr_methods = set()

        for result in all_results:
            if result is None:
                continue # This was a skipped image

            i = result['index']
            path_to_crop = result['path_to_crop']
            self.update_progress_report_batch(progress_report, i)

            # Aggregate failures
            if result['ocr_failed']: self.n_failed_OCR += 1
            if result['llm_failed']: self.n_failed_LLM_calls += 1

            # Aggregate costs and tokens
            self.OCR_cost += result.get('ocr_cost', 0)
            self.OCR_tokens_in += result.get('ocr_tokens_in', 0)
            self.OCR_tokens_out += result.get('ocr_tokens_out', 0)
            self.OCR_cost_in += result.get('ocr_cost_in', 0)
            self.OCR_cost_out += result.get('ocr_cost_out', 0)
            self.update_token_counters(result.get('nt_in', 0), result.get('nt_out', 0))
            if result.get('ocr_method'):
                collected_ocr_methods.add(result['ocr_method'])

            # Save individual file results
            final_JSON_response, final_WFO_record, final_GEO_record = self.update_final_response(
                result['response_candidate'],
                result['WFO_record'],
                result['GEO_record'],
                result['usage_report'],
                MODEL_NAME_FORMATTED,
                result['paths'],
                path_to_crop,
                result['nt_in'],
                result['nt_out']
            )
            
            # Update the GUI with the last processed item's result
            if json_report:
                json_report.set_JSON(final_JSON_response, final_WFO_record, final_GEO_record)

        # Clean up and finalize
        self.ocr_method = "|".join(sorted(list(collected_ocr_methods)))
        torch.cuda.empty_cache()
        gc.collect()

        self.update_progress_report_final(progress_report)
        if final_JSON_response:
            final_JSON_response = self.parse_final_json_response(final_JSON_response)

        return (final_JSON_response, final_WFO_record, final_GEO_record,
                self.total_tokens_in, self.total_tokens_out,
                self.OCR_cost, self.OCR_tokens_in, self.OCR_tokens_out,
                self.OCR_cost_in, self.OCR_cost_out, self.ocr_method)




    ##################################################################################################################################
    ################################################## LLM Helper Funcs ##############################################################
    ##################################################################################################################################
    def initialize_llm_model(self, cfg, logger, model_name, JSON_dict_structure, name_parts, is_azure=None, llm_object=None, config_vals_for_permutation=None):
        if 'LOCAL'in name_parts:
            if ('MIXTRAL' in name_parts) or ('MISTRAL' in name_parts):
                if 'CPU' in name_parts:
                    from vouchervision.LLM_local_cpu_MistralAI import LocalCPUMistralHandler
                    return LocalCPUMistralHandler(cfg, logger, model_name, JSON_dict_structure, config_vals_for_permutation)
                else:
                    from vouchervision.LLM_local_MistralAI import LocalMistralHandler
                    return LocalMistralHandler(cfg, logger, model_name, JSON_dict_structure, config_vals_for_permutation)
        elif "/" in ''.join(name_parts):
            from vouchervision.LLM_local_custom_fine_tune import LocalFineTuneHandler 
            return LocalFineTuneHandler(cfg, logger, model_name, JSON_dict_structure, config_vals_for_permutation)
        else:
            if 'PALM2' in name_parts:
                from vouchervision.LLM_GooglePalm2 import GooglePalm2Handler
                return GooglePalm2Handler(cfg, logger, model_name, JSON_dict_structure, config_vals_for_permutation)
            elif 'GEMINI' in name_parts:
                from vouchervision.LLM_GoogleGemini import GoogleGeminiHandler
                return GoogleGeminiHandler(cfg, logger, model_name, JSON_dict_structure, config_vals_for_permutation)
            elif 'MISTRAL' in name_parts and ('LOCAL' not in name_parts):
                from vouchervision.LLM_MistralAI import MistralHandler
                return MistralHandler(cfg, logger, model_name, JSON_dict_structure, config_vals_for_permutation)
            elif 'Hyperbolic' in name_parts:
                from vouchervision.LLM_Hyperbolic_Outlines import HyperbolicHandler
                return HyperbolicHandler(cfg, logger, model_name, JSON_dict_structure, config_vals_for_permutation)
            else:
                from vouchervision.LLM_OpenAI import OpenAIHandler
                return OpenAIHandler(cfg, logger, model_name, JSON_dict_structure, is_azure, llm_object, config_vals_for_permutation)

    def setup_prompt(self, ocr_text):
        """Generates a prompt using the provided OCR text."""
        Catalog = PromptCatalog()
        # Pass the ocr_text directly as a parameter
        prompt, _ = Catalog.prompt_SLTP(self.path_custom_prompts, OCR=ocr_text)
        return prompt
    
    def setup_JSON_dict_structure(self):
        Catalog = PromptCatalog()
        _, self.JSON_dict_structure = Catalog.prompt_SLTP(self.path_custom_prompts, OCR='Text')
    

    def initialize_token_counters(self):
        self.total_tokens_in = 0
        self.total_tokens_out = 0


    def update_progress_report_initial(self, progress_report):
        if progress_report is not None:
            progress_report.set_n_batches(len(self.img_paths))


    def update_progress_report_batch(self, progress_report, batch_index):
        if progress_report is not None:
            progress_report.update_batch(f"Working on image {batch_index + 1} of {len(self.img_paths)}")


    def should_skip_specimen(self, path_to_crop):
        return os.path.basename(path_to_crop) in self.completed_specimens


    def log_skipping_specimen(self, path_to_crop):
        self.logger.info(f'[Skipping] specimen {os.path.basename(path_to_crop)} already processed')

    
    def update_token_counters(self, nt_in, nt_out):
        self.total_tokens_in += nt_in
        self.total_tokens_out += nt_out


    def update_final_response(self, response_candidate, WFO_record, GEO_record, usage_report, MODEL_NAME_FORMATTED, paths, path_to_crop, nt_in, nt_out):
        filename_without_extension, txt_file_path, txt_file_path_OCR, txt_file_path_OCR_bounds, jpg_file_path_OCR_helper, json_file_path_wiki, txt_file_path_ind_prompt = paths
        # Saving the JSON and XLSX files with the response and updating the final JSON response
        if response_candidate is not None:
            final_JSON_response_updated = self.save_json_and_xlsx(self.Dirs, response_candidate, WFO_record, GEO_record, usage_report, MODEL_NAME_FORMATTED, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out)
            return final_JSON_response_updated, WFO_record, GEO_record
        else:
            final_JSON_response_updated = self.save_json_and_xlsx(self.Dirs, response_candidate, WFO_record, GEO_record, usage_report, MODEL_NAME_FORMATTED, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out)
            return final_JSON_response_updated, WFO_record, GEO_record


    def update_progress_report_final(self, progress_report):
        if progress_report is not None:
            progress_report.reset_batch("Batch Complete")


    def parse_final_json_response(self, final_JSON_response):
        try:
            return json.loads(final_JSON_response.strip('```').replace('json\n', '', 1).replace('json', '', 1))
        except:
            return final_JSON_response
    
    

    def generate_paths(self, path_to_crop, i):
        filename_without_extension = os.path.splitext(os.path.basename(path_to_crop))[0]
        txt_file_path = os.path.join(self.Dirs.transcription_ind, filename_without_extension + '.json')
        txt_file_path_OCR = os.path.join(self.Dirs.transcription_ind_OCR, filename_without_extension + '.json')
        txt_file_path_OCR_bounds = os.path.join(self.Dirs.transcription_ind_OCR_bounds, filename_without_extension + '.json')
        jpg_file_path_OCR_helper = os.path.join(self.Dirs.transcription_ind_OCR_helper, filename_without_extension + '.jpg')
        json_file_path_wiki = os.path.join(self.Dirs.transcription_ind_wiki, filename_without_extension + '.json')
        txt_file_path_ind_prompt = os.path.join(self.Dirs.transcription_ind_prompt, filename_without_extension + '.txt')

        self.logger.info(f'Working on {i+1}/{len(self.img_paths)} --- {filename_without_extension}')

        return filename_without_extension, txt_file_path, txt_file_path_OCR, txt_file_path_OCR_bounds, jpg_file_path_OCR_helper, json_file_path_wiki, txt_file_path_ind_prompt


    def save_json_and_xlsx(self, Dirs, response, WFO_record, GEO_record, usage_report, 
                           MODEL_NAME_FORMATTED, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out):
        if response is None:
            response = self.JSON_dict_structure
            # Insert 'filename' as the first key
            response = {'filename': filename_without_extension, **{k: v for k, v in response.items() if k != 'filename'}}
            self.write_json_to_file(txt_file_path, response)

            # Then add the null info to the spreadsheet
            response_null = self.create_null_row(filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper)
            self.add_data_to_excel_from_response(Dirs, self.path_transcription, response_null, WFO_record, GEO_record, usage_report, MODEL_NAME_FORMATTED, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in=0, nt_out=0)
        
        ### Set completed JSON
        else:
            response = self.clean_catalog_number(response, filename_without_extension)
            self.write_json_to_file(txt_file_path, response)
            # add to the xlsx file
            self.add_data_to_excel_from_response(Dirs, self.path_transcription, response, WFO_record, GEO_record, usage_report, MODEL_NAME_FORMATTED, filename_without_extension, path_to_crop, txt_file_path, jpg_file_path_OCR_helper, nt_in, nt_out)
        return response
    

    def process_specimen_batch(self, progress_report, json_report, is_real_run=False):
        if not self.has_key:
            self.logger.error(f'No API key found for {self.version_name}')
            raise Exception(f"No API key found for {self.version_name}")

        try:
            if is_real_run:
                progress_report.update_overall(f"Transcribing Labels")

            final_json_response, final_WFO_record, final_GEO_record, total_tokens_in, total_tokens_out, OCR_cost, OCR_tokens_in, OCR_tokens_out, OCR_cost_in, OCR_cost_out, OCR_method = self.send_to_LLM(self.is_azure, progress_report, json_report, self.model_name)
            
            return final_json_response, final_WFO_record, final_GEO_record, total_tokens_in, total_tokens_out, OCR_cost, OCR_tokens_in, OCR_tokens_out, OCR_cost_in, OCR_cost_out, OCR_method

        except Exception as e:
            self.logger.error(f"LLM call failed in process_specimen_batch: {e}")
            if progress_report is not None:
                progress_report.reset_batch(f"Batch Failed")
            self.close_logger_handlers()
            raise


    def close_logger_handlers(self):
        for handler in self.logger.handlers[:]:
            handler.close()
            self.logger.removeHandler(handler)


    # def process_specimen_batch_OCR_test(self, path_to_crop):
    #     for img_filename in os.listdir(path_to_crop):
    #         img_path = os.path.join(path_to_crop, img_filename)
    #     self.OCR, self.bounds, self.text_to_box_mapping = detect_text(img_path)



def space_saver(cfg, Dirs, logger):
    dir_out = cfg['leafmachine']['project']['dir_output']
    run_name = Dirs.run_name

    path_project = os.path.join(dir_out, run_name)

    if cfg['leafmachine']['project']['delete_temps_keep_VVE']:
        logger.name = '[DELETE TEMP FILES]'
        logger.info("Deleting temporary files. Keeping files required for VoucherVisionEditor.")
        delete_dirs = ['Archival_Components', 'Config_File']
        for d in delete_dirs:
            path_delete = os.path.join(path_project, d)
            if os.path.exists(path_delete):
                shutil.rmtree(path_delete)

    elif cfg['leafmachine']['project']['delete_all_temps']:
        logger.name = '[DELETE TEMP FILES]'
        logger.info("Deleting ALL temporary files!")
        delete_dirs = ['Archival_Components', 'Config_File', 'Original_Images', 'Cropped_Images']
        for d in delete_dirs:
            path_delete = os.path.join(path_project, d)
            if os.path.exists(path_delete):
                shutil.rmtree(path_delete)

        # Delete the transctiption folder, but keep the xlsx
        transcription_path = os.path.join(path_project, 'Transcription')
        if os.path.exists(transcription_path):
            for item in os.listdir(transcription_path):
                item_path = os.path.join(transcription_path, item)
                if os.path.isdir(item_path):  # if the item is a directory
                    if os.path.exists(item_path):
                        shutil.rmtree(item_path)  # delete the directory
